from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def load_compliance_rows(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["붙임1_법령준수여부"] if "붙임1_법령준수여부" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_label = str(row[0] or "").strip()
            result = normalize_result(row[1])
            item_no = item_label.split(".", 1)[0].strip()
            if item_no and result:
                rows[item_no] = result
        return rows
    finally:
        wb.close()


def normalize_result(value: object) -> str:
    text = str(value or "").strip()
    aliases = {
        "비대상": "해당없음",
        "근거부족": "확인필요",
    }
    return aliases.get(text, text)


def evaluate(predicted_path: Path, answer_path: Path) -> dict[str, object]:
    predicted = load_compliance_rows(predicted_path)
    answer = load_compliance_rows(answer_path)
    common_items = sorted(set(predicted).intersection(answer), key=lambda x: (len(x), x))
    rows = []
    correct = 0
    for item_no in common_items:
        pred = normalize_result(predicted[item_no])
        gold = normalize_result(answer[item_no])
        ok = pred == gold
        correct += int(ok)
        rows.append({"item_no": item_no, "predicted": pred, "answer": gold, "correct": ok})
    accuracy = correct / len(common_items) if common_items else 0.0
    return {
        "predicted_file": str(predicted_path),
        "answer_file": str(answer_path),
        "item_count": len(common_items),
        "correct_count": correct,
        "accuracy": round(accuracy, 4),
        "rows": rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="RFP 검토 결과를 정답 엑셀과 비교합니다.")
    parser.add_argument("--predicted", type=Path, required=True)
    parser.add_argument("--answer", type=Path, required=True)
    parser.add_argument("--json", type=Path)
    args = parser.parse_args()

    result = evaluate(args.predicted, args.answer)
    text = json.dumps(result, ensure_ascii=False, indent=2)
    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()
