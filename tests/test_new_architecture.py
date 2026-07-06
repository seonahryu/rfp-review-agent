import sqlite3
import tempfile
import unittest
from pathlib import Path


class NewArchitectureTests(unittest.TestCase):
    def test_rule_review_runs_without_llm(self):
        from agents.models import CandidatePage, RagContext
        from agents.rule_review_agent import RuleReviewAgent

        pages = [
            CandidatePage(
                page_no=1,
                page_text="과업심의위원회 심의를 완료하고 그 결과를 제안요청서에 반영한다.",
                text_length=32,
            )
        ]

        result = RuleReviewAgent().review("1", pages, RagContext(item_no="1"))

        self.assertEqual(result.item_no, "1")
        self.assertFalse(result.used_llm)
        self.assertEqual(result.result, "준수")
        self.assertGreaterEqual(result.confidence, 0.8)

    def test_rule_review_accepts_seventy_percent_keyword_coverage_from_rag(self):
        from agents.models import CandidatePage, RagContext, RagHit
        from agents.rule_review_agent import RuleReviewAgent

        rag = RagContext(
            item_no="19",
            hits=[
                RagHit(
                    item_no="19",
                    source_type="criteria_excel",
                    source_name="criteria.xlsx",
                    title="alpha beta gamma",
                    category="criteria",
                    snippet="alpha beta gamma delta epsilon",
                )
            ],
        )
        pages = [
            CandidatePage(
                page_no=5,
                page_text="alpha beta gamma delta are visible in this RFP page.",
                text_length=52,
            )
        ]

        result = RuleReviewAgent().review("19", pages, rag)

        self.assertEqual(result.result, "준수")
        self.assertIn("70%", result.reason)

    def test_table_and_attachment_agents_accept_seventy_percent_rag_coverage(self):
        from agents.attach_review_agent import AttachmentReviewAgent
        from agents.models import CandidatePage, RagContext, RagHit
        from agents.table_review_agent import TableReviewAgent

        rag = RagContext(
            item_no="11",
            hits=[
                RagHit(
                    item_no="11",
                    source_type="criteria_excel",
                    source_name="criteria.xlsx",
                    title="alpha beta gamma",
                    category="criteria",
                    snippet="alpha beta gamma delta epsilon",
                )
            ],
        )
        pages = [
            CandidatePage(
                page_no=7,
                page_text="alpha beta gamma delta appear in the evaluation table.",
                text_length=56,
                has_table_candidate=True,
                has_eval_table_candidate=True,
                has_attachment_candidate=True,
            )
        ]

        table_result = TableReviewAgent().review("11", pages, rag)
        attach_result = AttachmentReviewAgent().review("4", pages, rag)

        self.assertEqual(table_result.result, "준수")
        self.assertEqual(attach_result.result, "준수")

    def test_disabled_llm_vote_is_excluded_from_verification(self):
        from agents.models import ReviewResult
        from agents.verification_agent import VerificationAgent

        primary = ReviewResult(
            item_no="1",
            route_type="rule_review",
            result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[1],
            evidence_text=["evidence"],
            reason="rule evidence",
            recommendation="",
            needs_human_review=False,
            source="python_rule",
        )
        disabled = ReviewResult(
            item_no="1",
            route_type="llm_review",
            result="확인필요",
            is_target=None,
            confidence=0.4,
            evidence_pages=[1],
            evidence_text=["evidence"],
            reason="api disabled",
            recommendation="",
            needs_human_review=True,
            source="llm_disabled",
        )

        final = VerificationAgent().verify("1", [primary, disabled], parse_status="ok")

        self.assertEqual(final.final_status, "자동 확정 가능")
        self.assertEqual(final.final_result, "준수")

    def test_conflict_between_primary_and_low_model_escalates_to_strong_model(self):
        from agents.models import ReviewResult
        from agents.verification_agent import VerificationAgent

        class FakeLlm:
            def is_configured(self):
                return True

            def json_response(self, *args, **kwargs):
                return {
                    "result": "미준수",
                    "is_target": True,
                    "confidence": 0.91,
                    "evidence_pages": [99],
                    "evidence_text": ["hallucinated strong evidence"],
                    "reason": "escalated",
                    "recommendation": "보완 필요",
                    "needs_human_review": False,
                }

        first = ReviewResult(
            item_no="8",
            route_type="rule_review",
            result="준수",
            is_target=True,
            confidence=0.8,
            evidence_pages=[2],
            evidence_text=["rule evidence"],
            reason="rule",
            recommendation="",
            needs_human_review=False,
            source="python_rule",
        )
        second = ReviewResult(
            item_no="8",
            route_type="llm_review",
            result="미준수",
            is_target=True,
            confidence=0.82,
            evidence_pages=[3],
            evidence_text=["llm evidence"],
            reason="llm",
            recommendation="보완 필요",
            needs_human_review=False,
            source="openai_low",
            used_llm=True,
        )

        final = VerificationAgent(FakeLlm()).verify("8", [first, second], parse_status="ok")

        self.assertEqual(final.final_status, "상위 GPT 최종 판정")
        self.assertEqual(final.final_result, "미준수")
        self.assertEqual(final.evidence_pages, [3])
        self.assertEqual(final.evidence_text, ["llm evidence"])

    def test_orchestrator_accepts_existing_document_and_returns_summary(self):
        from orchestrator import RfpReviewPipeline

        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "rfp.db"
            output_dir = Path(tmp) / "out"
            conn = sqlite3.connect(db_path)
            conn.executescript(
                """
                CREATE TABLE rfp_document (
                    id INTEGER PRIMARY KEY,
                    document_name TEXT NOT NULL,
                    file_path TEXT,
                    total_pages INTEGER NOT NULL,
                    parse_status TEXT NOT NULL,
                    parse_warning_count INTEGER DEFAULT 0
                );
                CREATE TABLE rfp_page (
                    id INTEGER PRIMARY KEY,
                    document_id INTEGER NOT NULL,
                    page_no INTEGER NOT NULL,
                    page_text TEXT,
                    text_length INTEGER,
                    has_table_candidate INTEGER DEFAULT 0,
                    has_attachment_candidate INTEGER DEFAULT 0,
                    has_eval_table_candidate INTEGER DEFAULT 0,
                    has_toc_candidate INTEGER DEFAULT 0,
                    has_blind_candidate INTEGER DEFAULT 0,
                    has_commercial_sw_candidate INTEGER DEFAULT 0,
                    image_count INTEGER DEFAULT 0,
                    parser_warning TEXT
                );
                """
            )
            conn.execute("INSERT INTO rfp_document VALUES (1, 'sample.pdf', '', 1, 'ok', 0)")
            conn.execute(
                """
                INSERT INTO rfp_page (
                    id, document_id, page_no, page_text, text_length,
                    has_table_candidate, has_attachment_candidate, has_eval_table_candidate,
                    has_toc_candidate, has_blind_candidate, has_commercial_sw_candidate,
                    image_count, parser_warning
                )
                VALUES (1, 1, 1, ?, 32, 0, 0, 0, 0, 0, 0, 0, NULL)
                """,
                ("과업심의위원회 심의를 완료하고 그 결과를 제안요청서에 반영한다.",),
            )
            conn.commit()
            conn.close()

            summary = RfpReviewPipeline(db_path=db_path, output_dir=output_dir).review_existing_document(
                document_id=1,
                item_nos=["1"],
            )

            self.assertEqual(summary.document_id, 1)
            self.assertEqual(len(summary.final_reviews), 1)
            self.assertTrue(summary.excel_path.exists())

    def test_orchestrator_merges_item_two_one_into_item_two_final_output(self):
        from orchestrator import RfpReviewPipeline

        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "rfp.db"
            output_dir = Path(tmp) / "out"
            conn = sqlite3.connect(db_path)
            conn.executescript(
                """
                CREATE TABLE rfp_document (
                    id INTEGER PRIMARY KEY,
                    document_name TEXT NOT NULL,
                    file_path TEXT,
                    total_pages INTEGER NOT NULL,
                    parse_status TEXT NOT NULL,
                    parse_warning_count INTEGER DEFAULT 0
                );
                CREATE TABLE rfp_page (
                    id INTEGER PRIMARY KEY,
                    document_id INTEGER NOT NULL,
                    page_no INTEGER NOT NULL,
                    page_text TEXT,
                    text_length INTEGER,
                    has_table_candidate INTEGER DEFAULT 0,
                    has_attachment_candidate INTEGER DEFAULT 0,
                    has_eval_table_candidate INTEGER DEFAULT 0,
                    has_toc_candidate INTEGER DEFAULT 0,
                    has_blind_candidate INTEGER DEFAULT 0,
                    has_commercial_sw_candidate INTEGER DEFAULT 0,
                    image_count INTEGER DEFAULT 0,
                    parser_warning TEXT
                );
                """
            )
            conn.execute("INSERT INTO rfp_document VALUES (1, 'sample.pdf', '', 1, 'ok', 0)")
            conn.execute(
                """
                INSERT INTO rfp_page (
                    id, document_id, page_no, page_text, text_length,
                    has_table_candidate, has_attachment_candidate, has_eval_table_candidate,
                    has_toc_candidate, has_blind_candidate, has_commercial_sw_candidate,
                    image_count, parser_warning
                )
                VALUES (1, 1, 1, '상용SW 직접구매 BMT 직접구매 대상', 30, 0, 1, 0, 0, 0, 1, 0, NULL)
                """
            )
            conn.commit()
            conn.close()

            summary = RfpReviewPipeline(db_path=db_path, output_dir=output_dir).review_existing_document(
                document_id=1,
                item_nos=["2", "2-1"],
            )

            self.assertEqual([review.item_no for review in summary.final_reviews], ["2"])
            self.assertEqual(len(summary.final_reviews[0].reviews), 2)

    def test_item_two_merge_uses_target_subitem_when_present(self):
        from agents.models import FinalReview, ReviewResult
        from orchestrator import merge_final_reviews

        not_target_review = ReviewResult(
            item_no="2",
            route_type="attachment_review",
            result="해당없음",
            is_target=False,
            confidence=0.78,
            evidence_pages=[],
            evidence_text=[],
            reason="상용SW 직접구매 대상 근거 없음",
            recommendation="",
            needs_human_review=False,
            source="python_attachment_direct_purchase_not_target",
        )
        target_review = ReviewResult(
            item_no="2-1",
            route_type="attachment_review",
            result="미준수",
            is_target=True,
            confidence=0.84,
            evidence_pages=[20],
            evidence_text=["상용SW 직접구매 대상"],
            reason="대상 하위항목 근거 확인",
            recommendation="보완 필요",
            needs_human_review=False,
            source="openai_attachment",
        )
        merged = merge_final_reviews(
            [
                FinalReview(
                    item_no="2",
                    final_status="자동 확정 가능",
                    final_result=not_target_review.result,
                    is_target=not_target_review.is_target,
                    confidence=not_target_review.confidence,
                    evidence_pages=not_target_review.evidence_pages,
                    evidence_text=not_target_review.evidence_text,
                    reason=not_target_review.reason,
                    recommendation=not_target_review.recommendation,
                    reviews=[not_target_review],
                ),
                FinalReview(
                    item_no="2-1",
                    final_status="자동 확정 가능",
                    final_result=target_review.result,
                    is_target=target_review.is_target,
                    confidence=target_review.confidence,
                    evidence_pages=target_review.evidence_pages,
                    evidence_text=target_review.evidence_text,
                    reason=target_review.reason,
                    recommendation=target_review.recommendation,
                    reviews=[target_review],
                ),
            ]
        )

        self.assertEqual([review.item_no for review in merged], ["2"])
        self.assertEqual(merged[0].final_result, "미준수")
        self.assertTrue(merged[0].is_target)
        self.assertEqual(merged[0].evidence_pages, [20])

    def test_merge_final_reviews_keeps_evidence_page_text_pairs_aligned(self):
        from agents.models import FinalReview, ReviewResult
        from orchestrator import merge_final_reviews

        first_review = ReviewResult(
            item_no="17",
            route_type="attachment_review",
            result="준수",
            is_target=True,
            confidence=0.72,
            evidence_pages=[101],
            evidence_text=["101쪽 원문"],
            reason="first",
            recommendation="",
            needs_human_review=False,
            source="first",
        )
        second_review = ReviewResult(
            item_no="17",
            route_type="attachment_review",
            result="미준수",
            is_target=True,
            confidence=0.9,
            evidence_pages=[22],
            evidence_text=["22쪽 원문"],
            reason="second",
            recommendation="보완",
            needs_human_review=False,
            source="second",
        )

        merged = merge_final_reviews(
            [
                FinalReview(
                    item_no="17",
                    final_status="자동 확정 가능",
                    final_result=first_review.result,
                    is_target=first_review.is_target,
                    confidence=first_review.confidence,
                    evidence_pages=first_review.evidence_pages,
                    evidence_text=first_review.evidence_text,
                    reason=first_review.reason,
                    recommendation=first_review.recommendation,
                    reviews=[first_review],
                ),
                FinalReview(
                    item_no="17",
                    final_status="자동 확정 가능",
                    final_result=second_review.result,
                    is_target=second_review.is_target,
                    confidence=second_review.confidence,
                    evidence_pages=second_review.evidence_pages,
                    evidence_text=second_review.evidence_text,
                    reason=second_review.reason,
                    recommendation=second_review.recommendation,
                    reviews=[second_review],
                ),
            ]
        )

        self.assertEqual(merged[0].evidence_pages, [22, 101])
        self.assertEqual(merged[0].evidence_text, ["22쪽 원문", "101쪽 원문"])

    def test_verification_accepts_not_applicable_without_evidence_page(self):
        from agents.models import ReviewResult
        from agents.verification_agent import VerificationAgent

        review = ReviewResult(
            item_no="2",
            route_type="attachment_review",
            result="해당없음",
            is_target=False,
            confidence=0.78,
            evidence_pages=[],
            evidence_text=[],
            reason="RFP 본문에서 상용SW 직접구매 또는 분리발주 대상 사업으로 볼 근거가 확인되지 않았습니다.",
            recommendation="",
            needs_human_review=False,
            source="python_attachment_direct_purchase_not_target",
        )

        final = VerificationAgent().verify("2", [review], parse_status="ok")

        self.assertEqual(final.final_status, "자동 확정 가능")
        self.assertEqual(final.final_result, "해당없음")

    def test_item_fourteen_page_selection_prioritizes_requirement_summary(self):
        from agents.models import CandidatePage, RagContext, RagHit
        from orchestrator import select_evidence_pages

        pages = [
            CandidatePage(
                page_no=page_no,
                page_text="일반 문서 내용 소프트웨어 진흥법 요구사항 검토 기준",
                text_length=30,
            )
            for page_no in range(1, 13)
        ]
        pages.append(
            CandidatePage(
                page_no=20,
                page_text=(
                    "요구사항 총괄표 기능 요구사항 성능 요구사항 데이터 요구사항 "
                    "보안 요구사항 품질 요구사항 제약사항 프로젝트관리 요구사항"
                ),
                text_length=80,
            )
        )
        rag = RagContext(
            item_no="14",
            hits=[
                RagHit(
                    item_no="14",
                    source_type="criteria_db",
                    source_name="criteria",
                    title="요구사항 상세화",
                    category="명시",
                    snippet="요구사항 총괄표와 상세 요구사항 작성표를 활용",
                )
            ],
        )

        selected = select_evidence_pages("14", rag, pages, limit=5)

        self.assertEqual(selected[0].page_no, 20)

    def test_item_eleven_page_selection_prioritizes_ninety_ten_evaluation_page(self):
        from agents.models import CandidatePage, RagContext, RagHit
        from orchestrator import select_evidence_pages

        pages = [
            CandidatePage(
                page_no=page_no,
                page_text="일반 제안요청 내용 평가 기준",
                text_length=20,
            )
            for page_no in range(1, 13)
        ]
        pages.append(
            CandidatePage(
                page_no=33,
                page_text="평가원칙 - 기술능력평가점수(90점)와 가격평가점수(10점) 합산",
                text_length=50,
            )
        )
        rag = RagContext(
            item_no="11",
            hits=[
                RagHit(
                    item_no="11",
                    source_type="criteria_db",
                    source_name="criteria",
                    title="기술능력 평가비중",
                    category="명시",
                    snippet="기술능력평가의 평가비중을 90%로 명시",
                )
            ],
        )

        selected = select_evidence_pages("11", rag, pages, limit=5)

        self.assertEqual(selected[0].page_no, 33)

    def test_item_sixteen_page_selection_prioritizes_target_and_staff_pages(self):
        from agents.models import CandidatePage, RagContext, RagHit
        from orchestrator import select_evidence_pages

        pages = [
            CandidatePage(
                page_no=66,
                page_text="개인정보 처리 현황과 보안약점, 암호화 등 개인정보보호 관리 감독 사항",
                text_length=50,
            ),
            CandidatePage(
                page_no=9,
                page_text="본 사업의 적정 사업기간 산정을 위한 기능점수 규모는 726.4FP로 산정",
                text_length=50,
            ),
            CandidatePage(
                page_no=30,
                page_text="프로젝트 추진 일정에 따른 인력 투입 계획 제시 및 투입 인력 M/M 기준 50% 이상",
                text_length=70,
            ),
        ]
        rag = RagContext(
            item_no="16",
            hits=[
                RagHit(
                    item_no="16",
                    source_type="authoritative_criteria_db",
                    source_name="legal_item+legal_requirement",
                    title="투입인력 요구 및 관리 금지",
                    category="legal_review_criteria",
                    snippet="기능점수(FP) 방식으로 산정한 사업은 투입인력 요구 및 관리에 관한 내용 삭제",
                )
            ],
        )

        selected = select_evidence_pages("16", rag, pages, limit=3)

        self.assertEqual([page.page_no for page in selected[:2]], [30, 9])

    def test_item_thirteen_page_selection_prioritizes_proposal_compensation_pages(self):
        from agents.models import CandidatePage, RagContext, RagHit
        from orchestrator import select_evidence_pages

        pages = [
            CandidatePage(page_no=36, page_text="제안서 기술평가 항목 및 배점한도", text_length=30),
            CandidatePage(page_no=3, page_text="사업 예산: 450,317,164원", text_length=30),
            CandidatePage(page_no=35, page_text="본 사업은 20억원 미만의 SW사업으로 제안서 보상대상 사업에서 제외", text_length=60),
        ]
        rag = RagContext(
            item_no="13",
            hits=[
                RagHit(
                    item_no="13",
                    source_type="authoritative_criteria_db",
                    source_name="legal_item+legal_requirement",
                    title="SW사업 제안서 보상",
                    category="legal_review_criteria",
                    snippet="총사업금액 20억원 이상 SW개발사업. 제안서 보상 실시 여부 명시",
                )
            ],
        )

        selected = select_evidence_pages("13", rag, pages, limit=3)

        self.assertEqual(selected[0].page_no, 35)
        self.assertEqual(selected[1].page_no, 3)

    def test_basis_text_includes_reason_and_rfp_trigger_sentence(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import basis_text

        review = ReviewResult(
            item_no="6",
            route_type="rule_review",
            result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[25],
            evidence_text=["공급자는 지식재산권의 활용을 위하여 S/W 산출물의 반출을 요청할 수 있으며"],
            reason="규칙 기반 키워드가 RFP 본문에서 확인되었습니다.",
            recommendation="",
            needs_human_review=False,
            source="python_rule",
        )
        final = FinalReview(
            item_no="6",
            final_status="자동 확정 가능",
            final_result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[25],
            evidence_text=["공급자는 지식재산권의 활용을 위하여 S/W 산출물의 반출을 요청할 수 있으며"],
            reason="규칙 기반 키워드가 RFP 본문에서 확인되었습니다.",
            recommendation="",
            reviews=[review],
        )

        text = basis_text(final)

        self.assertIn("판단근거:", text)
        self.assertIn("RFP 트리거:", text)
        self.assertIn("p.25", text)
        self.assertIn("S/W 산출물의 반출", text)

    def test_item_fourteen_detail_comment_uses_page_and_below_wording(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import copy_paste_comment, item_label

        review = ReviewResult(
            item_no="14",
            route_type="rule_review",
            result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[10, 14, 15],
            evidence_text=["요구사항 총괄표", "요구사항 고유번호 SFR-001", "상세 세부내용"],
            reason="요구사항 총괄표와 상세 요구사항 작성 구조가 RFP 본문에서 확인되었습니다.",
            recommendation="",
            needs_human_review=False,
            source="python_rule_requirement_structure",
        )
        final = FinalReview(
            item_no="14",
            final_status="자동 확정 가능",
            final_result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[10, 14, 15],
            evidence_text=["요구사항 총괄표", "요구사항 고유번호 SFR-001", "상세 세부내용"],
            reason="요구사항 총괄표와 상세 요구사항 작성 구조가 RFP 본문에서 확인되었습니다.",
            recommendation="",
            reviews=[review],
        )

        self.assertIn("제안요청서 p.10 이하 명시", copy_paste_comment(final))
        self.assertIn("요구사항 상세화", item_label("14"))

    def test_item_fourteen_rule_evidence_starts_from_requirement_detail_not_toc_or_summary(self):
        from agents.models import CandidatePage, RagContext
        from agents.rule_review_agent import RuleReviewAgent

        pages = [
            CandidatePage(
                page_no=8,
                page_text="목차 요구사항 상세 내역 요구사항 고유번호 요구사항 명칭 정의 상세 세부내용 산출정보",
                text_length=50,
                rfp_printed_page_no=6,
                has_toc_candidate=True,
            ),
            CandidatePage(
                page_no=10,
                page_text=(
                    "요구사항 총괄표 기능 요구사항 성능 요구사항 데이터 요구사항 "
                    "보안 요구사항 품질 요구사항 제약사항 프로젝트관리 요구사항"
                ),
                text_length=80,
                rfp_printed_page_no=8,
            ),
            CandidatePage(
                page_no=14,
                page_text="요구사항 상세 내역 요구사항 고유번호 SFR-001 요구사항 명칭 정의 상세 세부내용 산출정보",
                text_length=70,
                rfp_printed_page_no=12,
            ),
        ]

        result = RuleReviewAgent().review("14", pages, RagContext(item_no="14"))

        self.assertEqual(result.evidence_pages, [12])

    def test_item_four_subcontract_evidence_keeps_only_explicit_project_disallow_page(self):
        from agents.models import CandidatePage, RagContext
        from agents.rule_review_agent import RuleReviewAgent

        pages = [
            CandidatePage(
                page_no=61,
                page_text="하도급 관련 법령 설명 및 불허 가능 사유 안내",
                text_length=30,
                rfp_printed_page_no=61,
            ),
            CandidatePage(
                page_no=62,
                page_text="본 사업은 소프트웨어 진흥법 제51조에 의거, 하도급은 원칙적으로 불허한다.",
                text_length=55,
                rfp_printed_page_no=62,
            ),
        ]

        result = RuleReviewAgent().review("4", pages, RagContext(item_no="4"))

        self.assertEqual(result.evidence_pages, [62])
        self.assertEqual(len(result.evidence_text), 1)

    def test_item_four_does_not_treat_resubcontract_limit_as_project_disallow(self):
        from agents.models import CandidatePage, RagContext
        from agents.rule_review_agent import RuleReviewAgent

        pages = [
            CandidatePage(
                page_no=62,
                page_text=(
                    "본 사업의 과업의 일부를 하도급하려는 경우 「소프트웨어 진흥법」 "
                    "제51조제1항에 따라 물품(상용소프트웨어 포함) 구매금액을 제외한 "
                    "소프트웨어사업금액의 100분의 50을 초과할 수 없으며, 같은 법 제3항에 "
                    "따라 다시 하도급은 원칙적으로 불허함."
                ),
                text_length=150,
                rfp_printed_page_no=62,
            )
        ]

        result = RuleReviewAgent().review("4", pages, RagContext(item_no="4"))

        self.assertNotEqual(result.result, "준수")
        self.assertNotEqual(result.source, "python_rule_subcontract_disallowed")

    def test_item_ten_direct_evidence_prefers_contract_method_wording(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=60,
                page_text="입찰 참가자격 및 사업 일반 안내",
                text_length=20,
                rfp_printed_page_no=60,
            ),
            CandidatePage(
                page_no=63,
                page_text=(
                    "제49조 제1항, 같은 법 시행령 제44조 제1항에 따라 "
                    "협상에 의한 계약체결 방법을 적용"
                ),
                text_length=70,
                rfp_printed_page_no=61,
            ),
        ]
        review = ReviewResult(
            item_no="10",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[60, 61],
            evidence_text=[pages[0].page_text, pages[1].page_text],
            reason="계약 방법 확인",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [61])
        self.assertIn("협상에 의한 계약체결", filtered.evidence_text[0])

    def test_item_nine_compliant_comment_says_no_specific_specification(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import basis_text, copy_paste_comment

        review = ReviewResult(
            item_no="9",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[2, 7, 13],
            evidence_text=["목차", "기능 요구사항", "산출정보"],
            reason="특정 상표, 모델명, 특정제품 기술용어가 발견되지 않았습니다.",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )
        final = FinalReview(
            item_no="9",
            final_status="자동 확정 가능",
            final_result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[2, 7, 13],
            evidence_text=["목차", "기능 요구사항", "산출정보"],
            reason="특정 상표, 모델명, 특정제품 기술용어가 발견되지 않았습니다.",
            recommendation="",
            reviews=[review],
        )

        self.assertEqual(copy_paste_comment(final), "특정규격 명시 없음")
        self.assertNotIn("RFP 트리거", basis_text(final))
        self.assertIn("특정규격 명시 없음", basis_text(final))

    def test_item_seventeen_noncompliant_comment_uses_legal_wording_without_toc_page(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import copy_paste_comment

        recommendation = (
            "국가기관 등의 장이 소프트웨어사업을 추진하는 경우「소프트웨어 진흥법」 제43조에 따라 "
            "민간시장에 미치는 영향을 분석하는 SW사업 영향평가를 실시하고, "
            "‘소프트웨어사업 계약 및 관리감독에 관한 지침’ 제5조에 따라 제안요청서에 "
            "별지 제1호서식 SW사업 영향평가 결과서를 작성하여 첨부하여야 합니다."
        )
        review = ReviewResult(
            item_no="17",
            route_type="attachment_review",
            result="미준수",
            is_target=True,
            confidence=0.82,
            evidence_pages=[2],
            evidence_text=["[붙임] 7. 소프트웨어사업 영향평가 검토결과서"],
            reason="목차 및 붙임 목록에는 표시되어 있으나 실제 첨부 본문이 확인되지 않았습니다.",
            recommendation=recommendation,
            needs_human_review=False,
            source="python_attachment_sw_impact_missing",
        )
        final = FinalReview(
            item_no="17",
            final_status="자동 확정 가능",
            final_result="미준수",
            is_target=True,
            confidence=0.82,
            evidence_pages=[2],
            evidence_text=review.evidence_text,
            reason=review.reason,
            recommendation=recommendation,
            reviews=[review],
        )

        comment = copy_paste_comment(final)

        self.assertEqual(comment, recommendation)
        self.assertNotIn("ㅇ (검토결과)", comment)
        self.assertNotIn("p.2", comment)

    def test_item_twelve_comment_uses_expected_recommendation(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import copy_paste_comment

        recommendation = (
            "제안요청서 p.33 차등점수제, pp.36-37 평가항목 및 배점한도 명시 "
            "-> “다만, 원점수 기준의 순위별 점수차가 차등점수보다 큰 경우에는 원점수차를 적용하며, "
            "차등점수를 부여한 후 기술능력평가점수와 가격평가점수를 합산하여 동점인 경우에는 "
            "기술능력평가점수에 따라 순위를 정함”을 추가 명시해주시기 바랍니다."
        )
        review = ReviewResult(
            item_no="12",
            route_type="table_review",
            result="보완필요",
            is_target=True,
            confidence=0.84,
            evidence_pages=[33, 36, 37],
            evidence_text=["차등점수제", "평가항목", "배점한도"],
            reason="차등점수제 세부 처리 문구가 누락되었습니다.",
            recommendation=recommendation,
            needs_human_review=False,
            source="python_table_differential_score_missing",
        )
        final = FinalReview(
            item_no="12",
            final_status="자동 확정 가능",
            final_result="보완필요",
            is_target=True,
            confidence=0.84,
            evidence_pages=[33, 36, 37],
            evidence_text=review.evidence_text,
            reason=review.reason,
            recommendation=recommendation,
            reviews=[review],
        )

        self.assertEqual(copy_paste_comment(final), recommendation)

    def test_format_pages_collapses_consecutive_ranges(self):
        from agents.report_agent import format_pages

        self.assertEqual(format_pages([32, 33, 34, 40]), "pp.32-34, 40")
        self.assertEqual(format_pages([24, 25]), "pp.24-25")

    def test_item_sixteen_report_uses_target_status_not_compliance_label(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import compliance_label, copy_paste_comment

        review = ReviewResult(
            item_no="16",
            route_type="llm_review",
            result="보완필요",
            is_target=True,
            confidence=0.8,
            evidence_pages=[3, 9],
            evidence_text=["사업 예산", "기능점수 726.4FP"],
            reason="기능점수 방식 산정 SW 개발사업",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )
        final = FinalReview(
            item_no="16",
            final_status="자동 확정 가능",
            final_result=review.result,
            is_target=True,
            confidence=0.8,
            evidence_pages=review.evidence_pages,
            evidence_text=review.evidence_text,
            reason=review.reason,
            recommendation="",
            reviews=[review],
        )

        self.assertEqual(compliance_label(final), "대상")
        self.assertEqual(copy_paste_comment(final), "제안요청서 pp.3, 9 기준 대상 사업으로 판단")

    def test_gpt_postprocess_keeps_only_direct_judgement_evidence(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import ReviewResult

        review = ReviewResult(
            item_no="13",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.85,
            evidence_pages=[3, 35],
            evidence_text=[
                "사업 예산: 450,317,164원",
                "제안서 보상 ○ 본 사업은 총 사업예산이 20억 미만으로 제안서 보상을 실시하지 않음",
            ],
            reason="대상 및 준수 판단",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review)

        self.assertEqual(filtered.evidence_pages, [35])

    def test_gpt_postprocess_uses_item_thirteen_direct_reason_page_even_when_result_not_applicable(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=3,
                page_text="사업명 정보시스템 보완 개발 사업 예산 450,317,164원\n- 3 -",
                text_length=50,
            ),
            CandidatePage(
                page_no=35,
                page_text="제안서 보상 ○ 본 사업은 총 사업예산이 20억 미만으로 제안서 보상을 실시하지 않음\n- 35 -",
                text_length=80,
                has_toc_candidate=True,
            ),
        ]
        review = ReviewResult(
            item_no="13",
            route_type="llm_review",
            result="해당없음",
            is_target=False,
            confidence=0.85,
            evidence_pages=[3],
            evidence_text=["사업 예산: 450,317,164원"],
            reason="대상 및 준수 판단",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [35])

    def test_gpt_postprocess_removes_target_only_evidence_for_all_items(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import RagContext, RagHit, ReviewResult

        rag = RagContext(
            item_no="5",
            hits=[
                RagHit(
                    item_no="5",
                    source_type="authoritative_criteria_db",
                    source_name="legal_item+legal_requirement",
                    title="SW사업 작업장소(원격개발)",
                    category="legal_review_criteria",
                    snippet="[REQUIRED_CRITERIA] 작업장소 상호협의 또는 제공여부 명시 원격지 개발 장소 보안요구사항",
                )
            ],
        )
        review = ReviewResult(
            item_no="5",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.85,
            evidence_pages=[3, 25],
            evidence_text=[
                "사업명: 정보시스템 보완 개발, 사업 예산: 450,317,164원",
                "S/W사업 수행을 위해 필요한 작업장소 등은 사업예산 내 계상되어 있으며 작업장소 등은 상호 협의하여 결정함",
            ],
            reason="대상 및 준수 판단",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, rag)

        self.assertEqual(filtered.evidence_pages, [25])

    def test_gpt_postprocess_rescans_item_five_direct_pages(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=3,
                page_text="사업명 정보시스템 보완 개발 사업 예산 450,317,164원\n- 3 -",
                text_length=50,
            ),
            CandidatePage(
                page_no=25,
                page_text=(
                    "S/W사업 수행을 위해 필요한 작업장소 등은 사업예산 내 계상되어 있으며 "
                    "작업장소 등은 상호 협의하여 결정함.\n- 25 -"
                ),
                text_length=100,
            ),
            CandidatePage(
                page_no=26,
                page_text=(
                    "공급자는 원격지 개발에 따른 개발방법 등에 대한 구체적인 방안을 제시하여야 하며 "
                    "원격지 개발 장소 보안요구사항 및 보안관리 대책을 준수하여야 함.\n- 26 -"
                ),
                text_length=120,
            ),
        ]
        review = ReviewResult(
            item_no="5",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.85,
            evidence_pages=[3],
            evidence_text=["사업명 정보시스템 보완 개발 사업 예산 450,317,164원"],
            reason="작업장소 상호협의, 비용 계상, 원격개발 장소 보안요구사항, 검토절차가 확인됨",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [25, 26])

    def test_gpt_postprocess_rescans_pages_when_model_outputs_only_target_page(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=3,
                page_text="사업명 정보시스템 보완 개발 사업 예산 450,317,164원\n- 3 -",
                text_length=50,
            ),
            CandidatePage(
                page_no=31,
                page_text=(
                    "PSR-002 하자보수 일반. 하자담보 책임기간은 본 사업이 종료한 날로부터 "
                    "1년으로 함.\n- 31 -"
                ),
                text_length=80,
            ),
        ]
        review = ReviewResult(
            item_no="8",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[3],
            evidence_text=["3페이지 사업예산 및 31페이지 하자담보 책임기간 1년 명시"],
            reason="대상 및 준수 판단",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [31])
        self.assertIn("하자담보", filtered.evidence_text[0])

    def test_item_eight_does_not_use_requirement_overview_as_direct_evidence(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=19,
                page_text=(
                    "요구사항 총괄표 프로젝트지원 PSR-001 일반 PSR-002 인수인계 "
                    "PSR-003 SW사업정보 제출 PSR-004 하자보수\n- 19 -"
                ),
                text_length=90,
            ),
            CandidatePage(
                page_no=103,
                page_text=(
                    "하자담보 책임 ○ 「(계약예규) 용역계약일반조건」 제58조 제2항과 제3항에 따라 "
                    "발주기관이 하자보수를 요청하면 계약상대자가 조치하여야 한다.\n- 103 -"
                ),
                text_length=120,
            ),
        ]
        review = ReviewResult(
            item_no="8",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[19],
            evidence_text=["PSR-004 하자보수"],
            reason="하자담보 책임기간이 1년 이내로 명시되어 준수",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [103])
        self.assertNotIn(19, filtered.evidence_pages)
        self.assertEqual(filtered.result, "보완필요")
        self.assertEqual(filtered.confidence, 0.86)
        self.assertIn("미작성", filtered.reason)
        self.assertNotIn("명시되어 준수", filtered.reason)
        self.assertIn("책임기간", filtered.recommendation)

    def test_item_eight_keeps_all_direct_evidence_needed_for_judgement(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=57,
                page_text=(
                    "하자담보 책임기간은 본 사업이 종료한 날로부터 1년으로 함.\n- 57 -"
                ),
                text_length=60,
            ),
            CandidatePage(
                page_no=91,
                page_text=(
                    "제안서 작성지침: 유지관리 계획, 조직, 절차, 범위 및 기간을 제시하여야 한다. "
                    "하자보수 계획, 조직, 절차, 범위 및 기간과 관련 활동을 제시하여야 한다.\n- 91 -"
                ),
                text_length=120,
            ),
            CandidatePage(
                page_no=97,
                page_text=(
                    "제안서 기술평가항목 및 배점표. 하자보수 계획의 적정성, 하자보수 절차의 적정성, "
                    "하자보수 범위 및 기간의 적정성을 평가한다.\n- 97 -"
                ),
                text_length=120,
            ),
            CandidatePage(
                page_no=103,
                page_text=(
                    "하자담보 책임 ○ 「(계약예규) 용역계약일반조건」 제58조 제2항과 제3항에 따라 "
                    "발주기관이 하자보수를 요청하면 계약상대자가 조치하여야 하며, "
                    "제58조 제2항 각 호의 경우는 유상 유지보수 또는 재개발로 본다. "
                    "계약상대자는 제58조 제3항 각 호의 사유로 발생한 하자에는 하자보수의 책임이 없음.\n- 103 -"
                ),
                text_length=180,
            ),
            CandidatePage(
                page_no=129,
                page_text=(
                    "서식 10 하도급 적정성 판단 자기평가표 ■ 소프트웨어사업 계약 및 관리감독에 관한 지침 "
                    "[별지 제8호서식] 하도급계약의 적정성 판단 자기평가표. "
                    "사업명 계약금액 하도급 계약기간 하자담보 책임기간 210mm×297mm 백상지 80g\n- 129 -"
                ),
                text_length=180,
            ),
            CandidatePage(
                page_no=87,
                page_text=(
                    "공동수급협정서 제13조(하자담보책임) 공동수급체는 공동수급체가 해산한 후 "
                    "해당 공사에 관하여 하자가 발생하였을 경우에는 연대하여 책임을 진다. "
                    "출자비율 구성원 운영위원회\n- 87 -"
                ),
                text_length=150,
            ),
        ]
        review = ReviewResult(
            item_no="8",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[57],
            evidence_text=["하자담보 책임기간은 1년으로 함"],
            reason="하자담보 책임기간 및 범위 준수",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [57, 103])
        self.assertEqual(filtered.result, "준수")

    def test_item_eight_keeps_multiple_direct_pages_even_when_same_aspect(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=54,
                page_text=(
                    "하자담보 책임기간은 본 사업이 종료한 날로부터 1년으로 함.\n- 54 -"
                ),
                text_length=60,
            ),
            CandidatePage(
                page_no=75,
                page_text=(
                    "하자담보 책임기간 명시. 본 사업에 포함된 산출물의 하자담보 책임기간은 "
                    "발주기관의 검사에 의하여 사업의 완성을 확인한 후 1년간으로 하며, "
                    "하자보수 범위는 개발시스템 및 도입 소프트웨어를 포함한 전체 시스템으로 함.\n- 75 -"
                ),
                text_length=160,
            ),
        ]
        review = ReviewResult(
            item_no="8",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.95,
            evidence_pages=[54],
            evidence_text=["하자담보 책임기간은 본 사업이 종료한 날로부터 1년으로 함"],
            reason="하자담보 책임기간 및 범위 준수",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [54, 75])
        self.assertEqual(filtered.result, "보완필요")
        self.assertIn("공통명시", filtered.reason)
        self.assertIn(
            "「(계약예규) 용역계약일반조건」 제58조 제2항과 제3항에 따라, 정한 기한내에 하자가 발생하여 발주기관이 하자보수를 계약상대자에게 요청한 경우 하자를 조치하여야 함",
            filtered.recommendation,
        )

    def test_item_eight_requires_common_statement_even_when_period_exists(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=54,
                page_text=(
                    "하자담보 책임기간은 본 사업이 종료한 날로부터 1년으로 함.\n- 54 -"
                ),
                text_length=60,
            )
        ]
        review = ReviewResult(
            item_no="8",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.95,
            evidence_pages=[54],
            evidence_text=["하자담보 책임기간은 본 사업이 종료한 날로부터 1년으로 함"],
            reason="하자담보 책임기간 준수",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.result, "보완필요")
        self.assertEqual(filtered.confidence, 0.95)
        self.assertIn("공통명시", filtered.reason)
        self.assertIn("item_8_missing_common_statement", filtered.warnings)
        self.assertFalse(filtered.needs_human_review)

    def test_item_eight_supplement_keeps_confidence_after_verification(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult
        from agents.verification_agent import VerificationAgent

        pages = [
            CandidatePage(
                page_no=54,
                page_text="하자담보 책임기간은 본 사업이 종료한 날로부터 1년으로 함.\n- 54 -",
                text_length=60,
            )
        ]
        review = ReviewResult(
            item_no="8",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.95,
            evidence_pages=[54],
            evidence_text=["하자담보 책임기간은 본 사업이 종료한 날로부터 1년으로 함"],
            reason="하자담보 책임기간 준수",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)
        final = VerificationAgent().verify("8", [filtered], parse_status="ok")

        self.assertEqual(final.final_result, "보완필요")
        self.assertEqual(final.final_status, "자동 확정 가능")
        self.assertEqual(final.confidence, 0.95)

    def test_direct_evidence_postprocess_does_not_cap_general_items_before_validation(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=page_no,
                page_text=(
                    "작업장소는 상호 협의하고 원격지 개발 방안을 제시하며 "
                    "보안요구사항을 준수한다.\n"
                    f"- {page_no} -"
                ),
                text_length=70,
            )
            for page_no in [21, 22, 23, 24]
        ]
        review = ReviewResult(
            item_no="5",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[21],
            evidence_text=["작업장소 상호협의"],
            reason="직접 근거 확인",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [21, 22, 23, 24])

    def test_item_five_recommendation_mentions_only_missing_requirements(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=75,
                page_text="2. 작업 장소 상호협의\n- 75 -",
                text_length=20,
            ),
            CandidatePage(
                page_no=76,
                page_text=(
                    "본 사업의 작업 장소는 원활한 사업수행을 위해 발주기관과 협의하여 결정함. "
                    "작업 장소 관련비용은 전체 사업예산에 계상되어 있으므로 제안사가 부담하여야 함.\n- 76 -"
                ),
                text_length=120,
            ),
        ]
        review = ReviewResult(
            item_no="5",
            route_type="llm_review",
            result="보완필요",
            is_target=True,
            confidence=0.9,
            evidence_pages=[75, 76],
            evidence_text=[
                "2. 작업 장소 상호협의",
                "작업 장소는 발주기관과 협의하여 결정하고 관련비용은 전체 사업예산에 계상되어 있음",
            ],
            reason="작업장소 협의와 비용 계상은 확인되나 일부 기준이 부족합니다.",
            recommendation=(
                "제안요청서에 작업장소 상호협의 또는 제공여부를 명확히 명시하고, "
                "작업장소 관련 비용 계상 여부를 포함해야 합니다. 또한, 공급자가 작업장소를 제시할 수 있는 절차와 "
                "발주기관의 검토 절차를 구체적으로 명시하고, 작업장소 관련 보안요구사항 및 원격개발에 따른 "
                "보안사고 대응방안도 포함하여야 합니다."
            ),
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertNotIn("작업장소 상호협의 또는 제공여부", filtered.recommendation)
        self.assertNotIn("비용 계상 여부", filtered.recommendation)
        self.assertIn("공급자가 작업장소를 제시할 수 있는 절차", filtered.recommendation)
        self.assertIn("원격개발에 따른 보안사고", filtered.recommendation)

    def test_item_nine_does_not_keep_whole_document_as_absence_evidence(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(page_no=3, page_text="사업 개요 및 시스템 현황\n- 3 -", text_length=20),
            CandidatePage(page_no=44, page_text="상용 SW 기반 도입을 위한 필요사항 제시\n- 44 -", text_length=30),
        ]
        review = ReviewResult(
            item_no="9",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.95,
            evidence_pages=[3, 44],
            evidence_text=["사업 개요 및 시스템 현황", "상용 SW 기반 도입을 위한 필요사항 제시"],
            reason="특정 상표, 모델명, 특정제품만 가능한 기능·성능에 대한 명시는 발견되지 않음",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [])
        self.assertEqual(filtered.evidence_text, [])

    def test_gpt_postprocess_corrects_document_page_index_to_rfp_printed_page_for_item_ten(self):
        from agents.gpt_judgement import keep_direct_judgement_evidence
        from agents.models import CandidatePage, ReviewResult

        pages = [
            CandidatePage(
                page_no=32,
                page_text=(
                    "2. 사업자 선정 방법 ○ 낙찰자 결정방법: 협상에 의한 계약 "
                    "협상에 의한 계약 체결 기준에 따라 협상적격자 등 선정 및 협상실시\n- 33 -"
                ),
                text_length=120,
            )
        ]
        review = ReviewResult(
            item_no="10",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[32],
            evidence_text=["사업자 선정 방법 낙찰자 결정방법: 협상에 의한 계약"],
            reason="준수 판단",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )

        filtered = keep_direct_judgement_evidence(review, pages=pages)

        self.assertEqual(filtered.evidence_pages, [33])

    def test_item_six_compliant_comment_does_not_list_each_required_statement(self):
        from agents.models import FinalReview
        from agents.report_agent import copy_paste_comment
        from agents.rule_review_agent import RuleReviewAgent
        from agents.models import CandidatePage, RagContext

        result = RuleReviewAgent().review(
            "6",
            [
                CandidatePage(
                    page_no=25,
                    page_text=(
                        "계약 목적물의 지식재산권은 발주기관과 계약상대자가 공동으로 소유한다. "
                        "공급자는 지식재산권의 활용을 위하여 S/W 산출물의 반출을 요청할 수 있으며 "
                        "누출금지정보로 명시한 정보를 삭제하고 활용하여야 하며 공급자 대표 명의의 확약서를 제출하여야 함. "
                        "반출된 S/W 산출물을 제3자에게 제공하려는 경우 발주기관의 사전승인을 받아야 함. "
                        "무단 유출하거나 누출금지정보를 삭제하지 않고 활용하는 경우 입찰 참가 자격을 제한함.\n- 23 -"
                    ),
                    text_length=260,
                )
            ],
            RagContext(item_no="6"),
        )
        final = FinalReview(
            item_no="6",
            final_status="자동 확정 가능",
            final_result=result.result,
            is_target=result.is_target,
            confidence=result.confidence,
            evidence_pages=result.evidence_pages,
            evidence_text=result.evidence_text,
            reason=result.reason,
            recommendation=result.recommendation,
            reviews=[result],
        )

        comment = copy_paste_comment(final)

        self.assertNotIn("ㅇ (검토결과)", comment)
        self.assertEqual(comment, "제안요청서 p.23 명시")
        self.assertNotIn("세부 명시 여부:", comment)
        self.assertNotIn("SW산출물 반출 요청절차", comment)
        self.assertNotIn("명시없음", comment)

    def test_item_five_compliant_comment_does_not_list_each_required_statement(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import copy_paste_comment

        review = ReviewResult(
            item_no="5",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[25],
            evidence_text=[
                (
                    "S/W사업 수행을 위해 필요한 작업장소 등은 사업예산 내 계상되어 있으므로 "
                    "관련 비용을 포함하여 제안가격을 산출하되 작업장소 등은 상호 협의하여 결정함. "
                    "공급자는 원격지 개발에 따른 개발방법 등에 대한 구체적인 방안을 제시하여야 하며 "
                    "원격지 보안관리 대책을 실시하여야 함."
                )
            ],
            reason="작업장소 상호협의, 원격지 개발 장소 제시, 보안요구사항이 확인되었습니다.",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )
        final = FinalReview(
            item_no="5",
            final_status="자동 확정 가능",
            final_result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[25],
            evidence_text=review.evidence_text,
            reason=review.reason,
            recommendation="",
            reviews=[review],
        )

        comment = copy_paste_comment(final)

        self.assertNotIn("ㅇ (검토결과)", comment)
        self.assertEqual(comment, "제안요청서 p.25 명시")
        self.assertNotIn("세부 명시 여부:", comment)
        self.assertNotIn("<원격지 개발 장소 보안요구사항>", comment)
        self.assertNotIn("명시없음", comment)

    def test_copy_paste_comment_excludes_target_only_pages_for_all_non_not_applicable_items(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import copy_paste_comment

        review = ReviewResult(
            item_no="13",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[3, 35],
            evidence_text=[
                "사업명 정보시스템 보완 개발 사업 예산 450,317,164원",
                "제안서 보상 ○ 본 사업은 총 사업예산이 20억 미만으로 제안서 보상을 실시하지 않음",
            ],
            reason="대상 및 준수 판단",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )
        final = FinalReview(
            item_no="13",
            final_status="자동 확정 가능",
            final_result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=review.evidence_pages,
            evidence_text=review.evidence_text,
            reason=review.reason,
            recommendation="",
            reviews=[review],
        )

        self.assertEqual(copy_paste_comment(final), "제안요청서 p.35 명시")

    def test_copy_paste_comment_keeps_item_three_direct_restriction_page(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import copy_paste_comment

        evidence = (
            "본 사업은 20억원 미만 사업으로 중소 소프트웨어사업자의 사업 참여 지원에 관한 지침에 따라 "
            "대기업 및 중견기업은 입찰에 참여할 수 없음"
        )
        review = ReviewResult(
            item_no="3",
            route_type="rule_review",
            result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[32],
            evidence_text=[evidence],
            reason="규칙 기반 핵심 근거가 RFP 본문에서 확인되었습니다.",
            recommendation="",
            needs_human_review=False,
            source="python_rule",
            used_llm=False,
        )
        final = FinalReview(
            item_no="3",
            final_status="자동 확정 가능",
            final_result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=review.evidence_pages,
            evidence_text=review.evidence_text,
            reason=review.reason,
            recommendation="",
            reviews=[review],
        )

        self.assertEqual(copy_paste_comment(final), "제안요청서 p.32 명시")

    def test_report_matches_manager_copy_paste_workbook_shape(self):
        from agents.models import FinalReview, ReviewResult
        from agents.report_agent import ReportAgent
        from openpyxl import load_workbook

        review = ReviewResult(
            item_no="1",
            route_type="rule_review",
            result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[10],
            evidence_text=["과업심의위원회 심의를 완료한다."],
            reason="필수 문구 확인",
            recommendation="",
            needs_human_review=False,
            source="python_rule",
        )
        final = FinalReview(
            item_no="1",
            final_status="자동 확정 가능",
            final_result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[10],
            evidence_text=["과업심의위원회 심의를 완료한다."],
            reason="필수 문구 확인",
            recommendation="",
            reviews=[review],
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = ReportAgent(Path(tmp)).write_excel(1, [final], [])
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    ["붙임1_법령준수여부", "별첨_항목별검토", "법제도_검토의견"],
                )
                self.assertEqual(
                    [wb["붙임1_법령준수여부"].cell(1, col).value for col in range(1, 5)],
                    ["검토항목", "법령준수 여부", "개선권고 관련 법적 근거", "사람 재검토"],
                )
                self.assertEqual(wb["별첨_항목별검토"].cell(1, 1).value, "항목")
                self.assertEqual(
                    wb["법제도_검토의견"].cell(1, 1).value,
                    "총 1개 항목 중 0개 항목 미준수 및 0개 항목 보완 권고",
                )
            finally:
                wb.close()

    def test_report_item_labels_match_current_legal_items(self):
        from agents.report_agent import item_label

        self.assertEqual(item_label("13"), "13. SW사업 제안서 보상")
        self.assertEqual(item_label("16"), "16. 투입인력 요구 및 관리 금지")
        self.assertEqual(item_label("17"), "17. SW사업 영향평가")

    def test_table_and_attachment_fallback_use_rfp_printed_page_number(self):
        from agents.attach_review_agent import AttachmentReviewAgent
        from agents.models import CandidatePage, RagContext
        from agents.table_review_agent import TableReviewAgent

        pages = [
            CandidatePage(
                page_no=40,
                page_text="관련 후보가 충분하지 않은 본문\n- 38 -",
                text_length=30,
                has_table_candidate=True,
                has_attachment_candidate=True,
            )
        ]

        table = TableReviewAgent().review("99", pages, RagContext(item_no="99"))
        attachment = AttachmentReviewAgent().review("99", pages, RagContext(item_no="99"))

        self.assertEqual(table.evidence_pages, [38])
        self.assertEqual(attachment.evidence_pages, [38])

    def test_rag_context_includes_authoritative_criteria_before_feedback_rows(self):
        from agents.rag_agent import RagAgent

        rag = RagAgent("rfp 법제도 검토항목.db")

        item_13 = rag.context_for_item("13")
        item_16 = rag.context_for_item("16")

        self.assertEqual(item_13.hits[0].source_type, "authoritative_criteria_db")
        self.assertIn("SW사업 제안서 보상", item_13.hits[0].title)
        self.assertIn("총사업금액 20억원 이상 SW개발사업", item_13.hits[0].snippet)
        self.assertIn("제안서 보상 실시 여부", item_13.hits[0].snippet)
        self.assertEqual(item_16.hits[0].source_type, "authoritative_criteria_db")
        self.assertIn("투입인력 요구 및 관리 금지", item_16.hits[0].title)
        self.assertIn("기능점수(FP)", item_16.hits[0].snippet)
        self.assertIn("투입인력 요구 및 관리에 관한 내용 삭제", item_16.hits[0].snippet)

    def test_gpt_prompt_contains_all_rfp_pages_not_top_n_candidates(self):
        import json

        from agents.gpt_judgement import build_user_prompt
        from agents.models import CandidatePage, RagContext

        pages = [
            CandidatePage(page_no=page_no, page_text=f"{page_no}쪽 본문", text_length=10)
            for page_no in range(1, 13)
        ]

        payload = json.loads(build_user_prompt("general", "16", pages, RagContext(item_no="16")))

        self.assertNotIn("candidate_pages", payload)
        self.assertEqual(len(payload["rfp_pages"]), 12)
        self.assertEqual(payload["rfp_pages"][-1]["document_page_index"], 12)

    def test_gpt_prompt_and_result_use_rfp_printed_page_number(self):
        import json

        from agents.gpt_judgement import build_user_prompt, normalize_evidence_pages
        from agents.models import CandidatePage, RagContext

        pages = [
            CandidatePage(page_no=32, page_text="계약 우선 적용 관련 문구\n- 33 -", text_length=30),
        ]

        payload = json.loads(build_user_prompt("general", "10", pages, RagContext(item_no="10")))

        self.assertEqual(payload["rfp_pages"][0]["document_page_index"], 32)
        self.assertEqual(payload["rfp_pages"][0]["rfp_printed_page_no"], 33)
        self.assertEqual(normalize_evidence_pages([32], pages), [33])

    def test_pipeline_normalizes_review_and_final_evidence_to_rfp_printed_page_number(self):
        from agents.models import CandidatePage, ReviewResult
        from agents.verification_agent import VerificationAgent
        from orchestrator import normalize_review_evidence_pages

        pages = [
            CandidatePage(
                page_no=93,
                page_text="중소 소프트웨어사업자의 사업참여 지원에 관한 지침 준수",
                text_length=32,
                rfp_printed_page_no=91,
            )
        ]
        review = ReviewResult(
            item_no="3",
            route_type="rule_review",
            result="준수",
            is_target=True,
            confidence=0.88,
            evidence_pages=[93],
            evidence_text=["중소 소프트웨어사업자의 사업참여 지원에 관한 지침 준수"],
            reason="규칙 기반 핵심 근거가 RFP 본문에서 확인되었습니다.",
            recommendation="",
            needs_human_review=False,
            source="python_rule",
            used_llm=False,
        )

        normalized = normalize_review_evidence_pages(review, pages)
        final = VerificationAgent().verify("3", [normalized], parse_status="ok")

        self.assertEqual(normalized.evidence_pages, [91])
        self.assertEqual(final.evidence_pages, [91])

    def test_keyword_evidence_handles_line_broken_words_and_late_page_text(self):
        from agents.models import CandidatePage
        from agents.rule_review_agent import evidence_for_keywords, evidence_window, loose_contains

        late_text = "앞부분 일반 안내\n" + ("일반 문장\n" * 40) + "본 사업은 차등점수\n제를 미적용한 사업임"

        self.assertTrue(loose_contains(late_text, "차등점수제"))
        self.assertIn("미적용한 사업임", evidence_window(late_text, "차등점수제", radius=80))

        matches = evidence_for_keywords(
            [CandidatePage(page_no=63, page_text=late_text, text_length=len(late_text))],
            ["차등점수제"],
        )

        self.assertEqual(matches[0][0], 63)
        self.assertIn("미적용한 사업임", matches[0][1])
        self.assertNotIn("앞부분 일반 안내", matches[0][1])

    def test_pipeline_uses_gpt_parser_for_new_pdf_runs(self):
        from agents.gpt_parser_agent import GptParserAgent
        from orchestrator import RfpReviewPipeline

        pipeline = RfpReviewPipeline(db_path=":memory:")

        self.assertIsInstance(pipeline.parser, GptParserAgent)

    def test_pipeline_passes_full_document_pages_to_primary_agent(self):
        import sqlite3
        import tempfile
        from pathlib import Path

        from agents.models import ReviewResult
        from orchestrator import RfpReviewPipeline

        class RecordingReview:
            def __init__(self):
                self.page_counts = []

            def review(self, item_no, pages, rag_context):
                self.page_counts.append(len(pages))
                return ReviewResult(
                    item_no=str(item_no),
                    route_type="llm_review",
                    result="해당없음",
                    is_target=False,
                    confidence=0.8,
                    evidence_pages=[],
                    evidence_text=[],
                    reason="recording",
                    recommendation="",
                    needs_human_review=False,
                    source="recording",
                )

        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "rfp.db"
            conn = sqlite3.connect(db_path)
            conn.executescript(
                """
                CREATE TABLE rfp_document (
                    id INTEGER PRIMARY KEY,
                    document_name TEXT NOT NULL,
                    file_path TEXT,
                    total_pages INTEGER NOT NULL,
                    parse_status TEXT NOT NULL,
                    parse_warning_count INTEGER DEFAULT 0
                );
                CREATE TABLE rfp_page (
                    id INTEGER PRIMARY KEY,
                    document_id INTEGER NOT NULL,
                    page_no INTEGER NOT NULL,
                    page_text TEXT,
                    text_length INTEGER,
                    has_table_candidate INTEGER DEFAULT 0,
                    has_attachment_candidate INTEGER DEFAULT 0,
                    has_eval_table_candidate INTEGER DEFAULT 0,
                    has_toc_candidate INTEGER DEFAULT 0,
                    has_blind_candidate INTEGER DEFAULT 0,
                    has_commercial_sw_candidate INTEGER DEFAULT 0,
                    image_count INTEGER DEFAULT 0,
                    parser_warning TEXT
                );
                CREATE TABLE legal_item (item_no TEXT, title TEXT, target_text TEXT);
                """
            )
            conn.execute("INSERT INTO rfp_document VALUES (1, 'sample.pdf', '', 12, 'ok', 0)")
            conn.execute("INSERT INTO legal_item VALUES ('16', '투입인력 요구 및 관리 금지', '')")
            for page_no in range(1, 13):
                conn.execute(
                    """
                    INSERT INTO rfp_page (
                        document_id, page_no, page_text, text_length,
                        has_table_candidate, has_attachment_candidate, has_eval_table_candidate,
                        has_toc_candidate, has_blind_candidate, has_commercial_sw_candidate,
                        image_count, parser_warning
                    )
                    VALUES (1, ?, ?, 10, 0, 0, 0, 0, 0, 0, 0, NULL)
                    """,
                    (page_no, f"{page_no}쪽 본문"),
                )
            conn.commit()
            conn.close()

            pipeline = RfpReviewPipeline(db_path=db_path, output_dir=Path(tmp) / "out")
            recorder = RecordingReview()
            pipeline.llm_review = recorder

            pipeline.review_existing_document(1, ["16"])

            self.assertEqual(recorder.page_counts, [12])


if __name__ == "__main__":
    unittest.main()
