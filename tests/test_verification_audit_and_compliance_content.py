import tempfile
import unittest
from pathlib import Path


class VerificationAuditAndComplianceContentTests(unittest.TestCase):
    def make_review(
        self,
        *,
        item_no="5",
        result="준수",
        is_target=True,
        pages=None,
        texts=None,
        reason="",
        recommendation="",
    ):
        from agents.models import FinalReview, ReviewResult

        pages = pages or []
        texts = texts or []
        review = ReviewResult(
            item_no=item_no,
            route_type="llm_review",
            result=result,
            is_target=is_target,
            confidence=0.9,
            evidence_pages=pages,
            evidence_text=texts,
            reason=reason,
            recommendation=recommendation,
            needs_human_review=False,
            source="test",
        )
        return FinalReview(
            item_no=item_no,
            final_status="자동 확정 가능",
            final_result=result,
            is_target=is_target,
            confidence=0.9,
            evidence_pages=pages,
            evidence_text=texts,
            reason=reason,
            recommendation=recommendation,
            reviews=[review],
        )

    def test_pass_two_audits_without_replacing_primary_result(self):
        from agents.models import CandidatePage, ReviewResult
        from agents.verification_agent import VerificationAuditAgent

        review = ReviewResult(
            item_no="5",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[25],
            evidence_text=["작업장소 상호협의 및 원격지 개발방법을 제시한다."],
            reason="작업장소, 원격지 개발방법, 보안관리 대책이 모두 확인됨",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
            used_llm=True,
        )
        pages = [
            CandidatePage(
                page_no=25,
                page_text="작업장소 상호협의 및 원격지 개발방법을 제시한다.",
                text_length=28,
            )
        ]

        audit = VerificationAuditAgent().audit_review(
            item_no="5",
            review=review,
            pages=pages,
            expected_route="llm",
            atomic_requirements=[
                {"id": "5-A", "text": "작업장소 상호협의"},
                {"id": "5-B", "text": "원격지 개발방법"},
                {"id": "5-C", "text": "보안관리 대책"},
            ],
        )

        self.assertFalse(audit.can_auto_accept)
        self.assertTrue(audit.requires_adjudication)
        self.assertEqual(review.result, "준수")
        self.assertIn("missing_required_requirement", [finding.finding_type for finding in audit.findings])

    def test_pass_three_adjudicator_requires_atomic_requirement_table(self):
        from agents.models import ReviewResult, VerificationAudit
        from agents.verification_agent import FinalAdjudicationAgent

        class FakeLlm:
            def is_configured(self):
                return True

            def json_response(self, *args, **kwargs):
                return {
                    "final_result": "보완필요",
                    "atomic_requirement_assessment": [
                        {"id": "5-A", "status": "met", "evidence_pages": [25], "evidence_text": ["작업장소 상호협의"], "reason": "확인됨"},
                        {"id": "5-B", "status": "met", "evidence_pages": [25], "evidence_text": ["원격지 개발방법"], "reason": "확인됨"},
                        {"id": "5-C", "status": "not_found", "evidence_pages": [], "evidence_text": [], "reason": "보안관리 대책 없음"},
                    ],
                    "reason": "필수 요건 일부 누락",
                    "recommendation": "보안관리 대책을 명시하시기 바랍니다.",
                    "evidence_pages": [25],
                    "evidence_text": ["작업장소 상호협의 및 원격지 개발방법"],
                    "confidence": 0.78,
                }

        primary = ReviewResult(
            item_no="5",
            route_type="llm_review",
            result="준수",
            is_target=True,
            confidence=0.86,
            evidence_pages=[25],
            evidence_text=["작업장소 상호협의 및 원격지 개발방법"],
            reason="모두 확인됨",
            recommendation="",
            needs_human_review=False,
            source="openai_general",
        )
        audit = VerificationAudit(
            item_no="5",
            audited_review_source="openai_general",
            can_auto_accept=False,
            requires_adjudication=True,
            audit_summary="필수 요건 일부 누락",
        )

        adjudicated = FinalAdjudicationAgent(FakeLlm()).adjudicate(
            item_no="5",
            primary_review=primary,
            audit=audit,
            atomic_requirements=[
                {"id": "5-A", "text": "작업장소 상호협의"},
                {"id": "5-B", "text": "원격지 개발방법"},
                {"id": "5-C", "text": "보안관리 대책"},
            ],
        )

        self.assertEqual(adjudicated.final_result, "보완필요")
        self.assertEqual(len(adjudicated.atomic_requirement_assessment), 3)
        self.assertEqual(adjudicated.atomic_requirement_assessment[2]["status"], "not_found")

    def test_compliant_content_is_simple_page_myeongsi(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            result="준수",
            pages=[25, 26],
            texts=["작업장소 상호협의 및 원격지 개발방법", "원격지 개발 보안관리 대책"],
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="5"))

        self.assertEqual(content.primary_evidence_pages, [25, 26])
        self.assertEqual(content.compliance_content, "제안요청서 pp.25-26 명시")

    def test_not_applicable_content_is_only_not_applicable(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(result="해당없음", is_target=False, pages=[30], texts=["대상 아님"])

        content = ComplianceContentAgent().generate(final, RagContext(item_no="8"))

        self.assertEqual(content.primary_evidence_pages, [])
        self.assertEqual(content.compliance_content, "해당없음")

    def test_recommendation_uses_only_direct_legal_review_pages(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="8",
            result="미준수",
            pages=[3, 57, 103],
            texts=[
                "사업개요 및 계약대상 확인",
                "SW 하자담보 책임기간을 1년으로 명시",
                "SW 하자담보 책임기간을 1년 이내로 명시",
            ],
            reason="SW 하자담보 책임기간 기준에 맞지 않음",
            recommendation="SW 하자담보 책임기간을 ‘1년’ 또는 ‘1년 이내’ 로 명시하시기 바랍니다.",
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="8"))

        self.assertEqual(content.primary_evidence_pages, [57, 103])
        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.57, p.103에 SW 하자담보 책임기간을 ‘1년’ 또는 ‘1년 이내’ 로 명시하시기 바랍니다.",
        )

    def test_partial_recommendation_uses_some_specified_pages(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="6",
            result="보완필요",
            pages=[50, 102],
            texts=[
                "산출물 활용 시 누출금지정보 삭제 일부 명시",
                "공급자 대표 명의 확약서 제출 일부 명시",
            ],
            recommendation="제안요청서에 ‘산출물 활용 시 누출금지정보 삭제와 공급자 대표 명의 확약서 제출’에 관하여 명시하시기 바랍니다.",
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="6"))

        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.50, p.102 일부 명시\n"
            "→ 제안요청서에 ‘산출물 활용 시 누출금지정보 삭제와 공급자 대표 명의 확약서 제출’에 관하여 명시하시기 바랍니다.",
        )

    def test_item_14_uses_below_first_evidence_page(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="14",
            result="준수",
            pages=[10, 11, 12, 13],
            texts=["세부요구사항 시작", "세부요구사항", "세부요구사항", "세부요구사항"],
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="14"))

        self.assertEqual(content.primary_evidence_pages, [10])
        self.assertEqual(content.compliance_content, "제안요청서 p.10 이하 명시")

    def test_report_agent_uses_generated_compliance_content_for_detail_comment(self):
        from agents.models import ComplianceContent
        from agents.report_agent import ReportAgent
        from openpyxl import load_workbook

        final = self.make_review(
            result="준수",
            pages=[25, 26],
            texts=["작업장소 상호협의", "보안관리 대책"],
        )
        final.compliance_content = ComplianceContent(
            item_no="5",
            content_type="compliance_statement",
            primary_evidence_pages=[25, 26],
            used_evidence_pages=[25, 26],
            compliance_content="제안요청서 pp.25-26 명시",
            tacit_knowledge_used=[],
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = ReportAgent(Path(tmp)).write_excel(1, [final], [])
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                detail = wb[wb.sheetnames[1]]
                self.assertEqual(detail.cell(2, 2).value, "제안요청서 pp.25-26 명시")
            finally:
                wb.close()


    def test_generic_recommendation_is_replaced_with_rag_requirement_detail(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext, RagHit

        final = self.make_review(
            item_no="6",
            result="미준수",
            pages=[102],
            texts=["산출물 활용 관련 일부 문구만 명시"],
            recommendation="관련 문구를 모두 명시해야 준수로 판단됩니다. 누락된 필수 문구를 보완하시기 바랍니다.",
        )
        rag = RagContext(
            item_no="6",
            hits=[
                RagHit(
                    item_no="6",
                    source_type="authoritative_criteria_db",
                    source_name="criteria",
                    title="SW사업 산출물 활용 보장",
                    category="legal_review_criteria",
                    snippet=(
                        "[CRITERION 1]\n"
                        "core_requirement_text: 산출물 활용 시 누출금지정보 삭제와 공급자 대표 명의 확약서 제출\n"
                        "requirement_text: 산출물 활용 시 누출금지정보 삭제와 공급자 대표 명의 확약서 제출을 명시"
                    ),
                )
            ],
        )

        content = ComplianceContentAgent().generate(final, rag)

        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.102 일부 명시\n"
            "→ 제안요청서에 ‘산출물 활용 시 누출금지정보 삭제와 공급자 대표 명의 확약서 제출’에 관하여 명시하시기 바랍니다.",
        )

    def test_example_page_numbers_are_not_copied_into_current_rfp_content(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="12",
            result="보완필요",
            pages=[51],
            texts=["차등점수제 적용 및 평가항목 배점표는 있으나 원점수 기준 점수차 처리 문구 없음"],
            recommendation=(
                "제안요청서 p.33 차등점수제, p.74 평가항목 및 배점한도 일부 명시 -> "
                "제안요청서에 ‘원점수 기준의 순위별 점수차가 차등점수보다 큰 경우에는 "
                "원점수차를 적용한다’는 내용을 명시하시기 바랍니다."
            ),
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="12", hits=[]))

        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.51 일부 명시\n"
            "→ 제안요청서에 ‘원점수 기준의 순위별 점수차가 차등점수보다 큰 경우에는 원점수차를 적용한다’는 내용을 명시하시기 바랍니다.",
        )
        self.assertNotIn("p.33", content.compliance_content)
        self.assertNotIn("p.74", content.compliance_content)

    def test_partial_content_separates_existing_statement_from_action(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="17",
            result="미준수",
            pages=[104],
            texts=["소프트웨어사업 영향평가를 미리 실시한 사업임. 별지 제1호 서식에 따른 결과서를 첨부(별첨 2 참조)"],
            reason="소프트웨어사업 영향평가 실시 명시는 있으나, 결과서의 실제 검토 내용이 확인되지 않습니다.",
            recommendation=(
                "국가기관 등의 장이 소프트웨어사업을 추진하는 경우「소프트웨어 진흥법」 제43조에 따라 "
                "민간시장에 미치는 영향을 분석하는 SW사업 영향평가를 실시하고, ‘소프트웨어사업 계약 및 "
                "관리감독에 관한 지침’ 제5조에 따라 제안요청서에 별지 제1호서식 SW사업 영향평가 결과서를 "
                "작성하여 첨부하여야 합니다."
            ),
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="17"))

        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.104에 SW사업 영향평가 결과서를 작성하여 첨부하시기 바랍니다.",
        )
        self.assertNotIn("소프트웨어사업 영향평가를 미리 실시한 사업임", content.compliance_content)

    def test_non_substantive_form_pages_are_excluded_from_content_pages(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="1",
            result="보완필요",
            pages=[102, 120],
            texts=[
                "과업심의위원회 개최 여부 일부 명시",
                "서식 1. 과업심의위원회 관련 빈 양식",
            ],
            recommendation="제안요청서에 과업심의위원회 개최 여부를 명시하시기 바랍니다.",
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="1"))

        self.assertEqual(content.primary_evidence_pages, [102])
        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.102 일부 명시\n"
            "→ 제안요청서에 과업심의위원회 개최 여부를 명시하시기 바랍니다.",
        )

    def test_summary_list_pages_are_excluded_from_content_pages(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="18",
            result="보완필요",
            pages=[19, 104],
            texts=[
                "요구사항 총괄표 목록",
                "소프트웨어사업정보 제출 일부 명시",
            ],
            recommendation="제안요청서에 소프트웨어사업정보 제출 절차를 명시하시기 바랍니다.",
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="18"))

        self.assertEqual(content.primary_evidence_pages, [104])
        self.assertNotIn("p.19", content.compliance_content)
        self.assertIn("소프트웨어사업정보 제출", content.compliance_content)

    def test_requirement_id_overview_pages_are_excluded_from_content_pages(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="18",
            result="준수",
            pages=[19, 57],
            texts=[
                "요구사항 분류 프로젝트지원 PSR-001 인수인계 PSR-002 일반 PSR-003 SW사업정보 저장소 데이터 작성 및 제출 PSR-004 하자보수",
                "요구사항 분류 프로젝트지원 요구사항 고유번호 PSR-003 요구사항 명칭 SW사업정보 저장소 데이터 작성 및 제출",
            ],
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="18"))

        self.assertEqual(content.primary_evidence_pages, [57])
        self.assertEqual(content.compliance_content, "제안요청서 p.57 명시")

    def test_attachment_title_only_pages_are_excluded_from_content_pages(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="17",
            result="미준수",
            pages=[104, 111],
            texts=[
                "소프트웨어사업 영향평가를 미리 실시한 사업임. 별지 제1호 서식에 따른 소프트웨어사업 영향평가 결과서를 첨부",
                "별첨 2 소프트웨어사업 영향평가 검토결과서 111",
            ],
            recommendation="국가기관 등의 장이 소프트웨어사업을 추진하는 경우 SW사업 영향평가 결과서를 작성하여 첨부하여야 합니다.",
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="17"))

        self.assertEqual(content.primary_evidence_pages, [104])
        self.assertNotIn("p.111", content.compliance_content)
        self.assertNotIn("일부 명시", content.compliance_content)
        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.104에 SW사업 영향평가 결과서를 작성하여 첨부하시기 바랍니다.",
        )

    def test_title_only_attachment_gap_is_not_partial_content(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="15",
            result="미준수",
            pages=[103],
            texts=["소프트웨어 개발사업의 적정 사업기간 산정 기준 적용 명시"],
            reason="적정 사업기간 종합산정서 제목만 있고 실제 산정 내용이 없는 빈 문서로 확인됩니다.",
            recommendation=(
                "소프트웨어사업 계약 및 관리감독에 관한 지침 제10조에 의거 별지 제4호서식 "
                "소프트웨어 개발사업의 적정 사업기간 종합 산정서(단, 위원명 및 서명은 제외한다)를 첨부하시기 바랍니다."
            ),
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="15"))

        self.assertNotIn("일부 명시", content.compliance_content)
        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.103에 소프트웨어사업 계약 및 관리감독에 관한 지침 제10조에 의거 별지 제4호서식 "
            "소프트웨어 개발사업의 적정 사업기간 종합 산정서(단, 위원명 및 서명은 제외한다)를 첨부하시기 바랍니다.",
        )

    def test_missing_noncompliant_item_without_evidence_uses_recommendation(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="15",
            result="미준수",
            pages=[],
            texts=[],
            reason="소프트웨어 개발사업의 적정 사업기간 종합 산정서 첨부가 확인되지 않았습니다.",
            recommendation=(
                "소프트웨어사업 계약 및 관리감독에 관한 지침 제10조에 의거 별지 제4호서식 "
                "소프트웨어 개발사업의 적정 사업기간 종합 산정서(단, 위원명 및 서명은 제외한다)를 첨부하시기 바랍니다."
            ),
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="15"))

        self.assertEqual(content.primary_evidence_pages, [])
        self.assertEqual(
            content.compliance_content,
            "소프트웨어사업 계약 및 관리감독에 관한 지침 제10조에 의거 별지 제4호서식 "
            "소프트웨어 개발사업의 적정 사업기간 종합 산정서(단, 위원명 및 서명은 제외한다)를 첨부하시기 바랍니다.",
        )
        self.assertEqual(content.warnings, [])

    def test_page_prefix_does_not_duplicate_rfp_subject(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="17",
            result="미준수",
            pages=[104],
            texts=["SW사업 영향평가 일부 내용만 명시"],
            recommendation=(
                "제안요청서에 국가기관 등의 장이 소프트웨어사업을 추진하는 경우 "
                "SW사업 영향평가 결과서를 작성하여 첨부하여야 합니다."
            ),
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="17"))

        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.104 일부 명시\n"
            "→ SW사업 영향평가 결과서를 작성하여 첨부하시기 바랍니다.",
        )
        self.assertNotIn("p.104에 제안요청서에", content.compliance_content)

    def test_page_prefix_omits_particle_before_section_target(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="12",
            result="미준수",
            pages=[95, 96, 97],
            texts=[
                "제안서 기술평가항목 및 배점표",
                "제안서 기술평가항목 및 배점표",
                "제안서 기술평가항목 및 배점표",
            ],
            recommendation="제안서 기술평가항목 및 배점표에 하도급계획 적정성 평가항목을 5점 이상으로 명시해주시기 바랍니다.",
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="12"))

        self.assertEqual(
            content.compliance_content,
            "제안요청서 pp.95-97 제안서 기술평가항목 및 배점표에 하도급계획 적정성 평가항목을 5점 이상으로 명시해주시기 바랍니다.",
        )

    def test_rag_requirement_is_reduced_to_actionable_core_phrase(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext, RagHit

        final = self.make_review(
            item_no="7",
            result="미준수",
            pages=[102],
            texts=["지식재산권 귀속 일부 내용만 확인됨"],
            recommendation="관련 문구를 모두 명시해야 준수로 판단됩니다. 누락된 필수 문구를 보완하시기 바랍니다.",
        )
        rag = RagContext(
            item_no="7",
            hits=[
                RagHit(
                    item_no="7",
                    source_type="authoritative_criteria_db",
                    source_name="criteria",
                    title="개발SW의 공동활용 사전명시",
                    category="legal_review_criteria",
                    snippet=(
                        "core_requirement_text: 제안요청서 등에 지식재산의 공동귀속 적용 여부 명시 "
                        "※ 발주기관 귀속의 경우, 계약목적물의 특수성(국방, 외교관계, 국가안전보장 등)에 따른 구체적 사유"
                    ),
                )
            ],
        )

        content = ComplianceContentAgent().generate(final, rag)

        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.102 일부 명시\n"
            "→ 제안요청서에 ‘지식재산의 공동귀속 적용 여부’에 관하여 명시하시기 바랍니다.",
        )
        self.assertNotIn("※", content.compliance_content)

    def test_partial_content_never_copies_evidence_fragment(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext, RagHit

        final = self.make_review(
            item_no="7",
            result="미준수",
            pages=[102],
            texts=[
                "한 경우 거부할 수 있으며, 공급자가 유효한 정보보호 체계 인증 또는 소프트웨어프로세스 품질인증을 보유한 경우"
            ],
            reason="지식재산의 공동귀속 적용 여부는 일부 관련 조항만 확인됩니다.",
            recommendation="관련 문구를 모두 명시해야 준수로 판단됩니다. 누락된 필수 문구를 보완하시기 바랍니다.",
        )
        rag = RagContext(
            item_no="7",
            hits=[
                RagHit(
                    item_no="7",
                    source_type="authoritative_criteria_db",
                    source_name="criteria",
                    title="개발SW의 공동활용 사전명시",
                    category="legal_review_criteria",
                    snippet="core_requirement_text: 제안요청서 등에 지식재산의 공동귀속 적용 여부 명시",
                )
            ],
        )

        content = ComplianceContentAgent().generate(final, rag)

        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.102 일부 명시\n"
            "→ 제안요청서에 ‘지식재산의 공동귀속 적용 여부’에 관하여 명시하시기 바랍니다.",
        )
        self.assertNotIn("정보보호 체계 인증", content.compliance_content)

    def test_quoted_requirement_drops_rfp_etc_prefix(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="7",
            result="미준수",
            pages=[102],
            texts=["지식재산권 귀속 일부 명시"],
            reason="지식재산의 공동귀속 적용 여부는 일부 관련 조항만 확인됩니다.",
            recommendation="제안요청서에 ‘제안요청서 등에 지식재산의 공동귀속 적용 여부’에 관하여 명시하시기 바랍니다.",
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="7"))

        self.assertEqual(
            content.compliance_content,
            "제안요청서 p.102 일부 명시\n"
            "→ 제안요청서에 ‘지식재산의 공동귀속 적용 여부’에 관하여 명시하시기 바랍니다.",
        )

    def test_partial_action_removes_example_explanation(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="5",
            result="보완필요",
            pages=[75, 76],
            texts=["작업장소 상호협의 일부 명시", "작업장소 보안요구사항 일부 명시"],
            reason="작업장소 관련 비용 계상 여부가 미흡합니다.",
            recommendation=(
                "제안요청서에 작업장소 관련 비용 계상 여부를 구체적으로 포함해야 합니다. "
                "예를 들어, 작업장소 비용은 사업예산 내 계상되어 제안가격에 포함됨을 명확히 하고, "
                "보안요구사항에 따른 위험요인 식별 및 대응방안 제시를 요구하는 내용을 추가해야 합니다."
            ),
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="5"))

        self.assertEqual(
            content.compliance_content,
            "제안요청서 pp.75-76 일부 명시\n"
            "→ 제안요청서에 작업장소 관련 비용 계상 여부를 구체적으로 포함해야 합니다.",
        )
        self.assertNotIn("예를 들어", content.compliance_content)

    def test_item_nine_compliant_content_does_not_list_absence_evidence_pages(self):
        from agents.compliance_content_agent import ComplianceContentAgent
        from agents.models import RagContext

        final = self.make_review(
            item_no="9",
            result="준수",
            pages=[3, 6, 15, 16, 17, 18],
            texts=["사업 개요", "시스템 현황", "요구사항", "보안대책", "일반 내용", "일반 내용"],
            reason="RFP 내 특정 상표, 모델명, 특정제품만 가능한 기능·성능에 대한 명시는 발견되지 않음",
        )

        content = ComplianceContentAgent().generate(final, RagContext(item_no="9"))

        self.assertEqual(content.primary_evidence_pages, [])
        self.assertEqual(content.compliance_content, "특정규격 명시 없음")


if __name__ == "__main__":
    unittest.main()
