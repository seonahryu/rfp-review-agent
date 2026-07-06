from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

from agents.models import RagContext, RagHit


TOKEN_RE = re.compile(r"[0-9A-Za-z가-힣]{2,}")
ITEM_ALIASES = {
    "2": ["2", "2-1"],
    "2-1": ["2", "2-1"],
}


class RagAgent:
    def __init__(self, db_path: Path | str) -> None:
        self.db_path = Path(db_path)

    def ensure_schema(self) -> None:
        conn = connect(self.db_path)
        try:
            ensure_rag_schema(conn)
        finally:
            conn.close()

    def rebuild_from_excels(self, excel_paths: Iterable[Path]) -> int:
        conn = connect(self.db_path)
        ensure_rag_schema(conn)
        conn.execute("DELETE FROM rag_fts")
        conn.execute("DELETE FROM rag_document")
        count = 0
        for path in excel_paths:
            if path.name.startswith("~$"):
                continue
            for doc in documents_from_excel(path):
                insert_rag_document(conn, doc)
                count += 1
        conn.commit()
        conn.close()
        return count

    def context_for_item(self, item_no: str, query: str = "", limit: int = 8) -> RagContext:
        conn = connect(self.db_path)
        ensure_rag_schema(conn)
        try:
            authoritative_hits = load_authoritative_criteria_hits(conn, item_no)
            rows = search(conn, query or default_query_for_item(item_no), item_no=item_no, limit=limit)
            if not rows:
                rows = item_document_fallback(conn, item_no=item_no, limit=limit)
            rows = rows[:context_hit_limit(item_no)]
            hits = [
                RagHit(
                    item_no=row["item_no"],
                    source_type=row["source_type"],
                    source_name=row["source_name"],
                    title=row["title"] or "",
                    category=row["category"],
                    snippet=row["snippet"] or "",
                    page_or_row=format_location(row),
                    score=float(row["rank"]) if row["rank"] is not None else None,
                )
                for row in rows
            ]
            hits = merge_rag_hits(authoritative_hits, hits)
            hits.extend(load_tacit_knowledge_hits(conn, item_no))
            return RagContext(item_no=item_no, hits=hits)
        finally:
            conn.close()


def connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_rag_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS rag_document (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_type TEXT NOT NULL,
            source_name TEXT NOT NULL,
            sheet_name TEXT,
            row_no INTEGER,
            item_no TEXT,
            title TEXT,
            category TEXT,
            content TEXT NOT NULL,
            metadata_json TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE VIRTUAL TABLE IF NOT EXISTS rag_fts USING fts5(
            title,
            category,
            content,
            item_no UNINDEXED,
            source_type UNINDEXED,
            source_name UNINDEXED,
            sheet_name UNINDEXED,
            row_no UNINDEXED,
            doc_id UNINDEXED,
            tokenize = 'unicode61'
        );
        CREATE TABLE IF NOT EXISTS legal_tacit_knowledge (
            item_no TEXT NOT NULL,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    ensure_tacit_knowledge_schema(conn)
    conn.commit()


def documents_from_legal_tables(conn: sqlite3.Connection) -> list[dict[str, object]]:
    docs: list[dict[str, object]] = []
    requirements_by_item = load_requirements_by_item(conn)
    reference_examples_by_item = load_reference_examples_by_item(conn)
    item_titles = load_item_titles(conn)

    if table_exists(conn, "legal_item"):
        for row in conn.execute("SELECT item_no, title, target_text FROM legal_item ORDER BY item_no"):
            requirements = requirements_by_item.get(row["item_no"], [])
            reference_examples = reference_examples_by_item.get(row["item_no"], [])
            docs.append(
                {
                    "source_type": "criteria_db",
                    "source_name": source_name_for_criteria(reference_examples),
                    "item_no": row["item_no"],
                    "title": row["title"],
                    "category": "legal_review_criteria",
                    "content": build_enriched_criteria_content(row, requirements, reference_examples),
                    "metadata": {
                        "tables": criteria_source_tables(reference_examples),
                        "requirement_count": len(requirements),
                        "reference_example_count": len(reference_examples),
                    },
                }
            )

    for item_no, requirements in requirements_by_item.items():
        item_title = item_titles.get(item_no)
        for idx, requirement in enumerate(requirements, start=1):
            docs.append(
                {
                    "source_type": "requirement_db",
                    "source_name": "legal_requirement",
                    "row_no": idx,
                    "item_no": item_no,
                    "title": make_requirement_title(
                        item_no,
                        item_title,
                        requirement["category"],
                        requirement["requirement_text"],
                    ),
                    "category": requirement["category"] or "criteria",
                    "content": build_requirement_content(requirement),
                    "metadata": {"table": "legal_requirement"},
                }
            )

    if table_exists(conn, "legal_reference_example"):
        for idx, row in enumerate(
            conn.execute(
                "SELECT item_no, reference_type, reference_subtype, content "
                "FROM legal_reference_example"
            ),
            start=1,
        ):
            docs.append(
                {
                    "source_type": "reference_example_db",
                    "source_name": "legal_reference_example",
                    "row_no": idx,
                    "item_no": row["item_no"],
                    "title": row["reference_type"] or f"{row['item_no']} reference example",
                    "category": row["reference_subtype"] or "reference_example",
                    "content": row["content"],
                    "metadata": {"table": "legal_reference_example"},
                }
            )

    return docs


def load_requirements_by_item(conn: sqlite3.Connection) -> dict[str, list[dict[str, str]]]:
    if not table_exists(conn, "legal_requirement"):
        return {}
    requirements: dict[str, list[dict[str, str]]] = {}
    rows = conn.execute(
        'SELECT item_no, category, requirement_text, example_sentence, "case" AS case_text '
        "FROM legal_requirement ORDER BY item_no"
    )
    for row in rows:
        requirements.setdefault(row["item_no"], []).append(
            {
                "category": row["category"] or "",
                "requirement_text": row["requirement_text"] or "",
                "example_sentence": row["example_sentence"] or "",
                "case_text": row["case_text"] or "",
            }
        )
    return requirements


def load_reference_examples_by_item(conn: sqlite3.Connection) -> dict[str, list[dict[str, str]]]:
    if not table_exists(conn, "legal_reference_example"):
        return {}
    examples: dict[str, list[dict[str, str]]] = {}
    rows = conn.execute(
        "SELECT item_no, reference_type, reference_subtype, content "
        "FROM legal_reference_example ORDER BY item_no"
    )
    for row in rows:
        examples.setdefault(row["item_no"], []).append(
            {
                "reference_type": row["reference_type"] or "",
                "reference_subtype": row["reference_subtype"] or "",
                "content": row["content"] or "",
            }
        )
    return examples


def load_item_titles(conn: sqlite3.Connection) -> dict[str, str]:
    if not table_exists(conn, "legal_item"):
        return {}
    return {
        row["item_no"]: row["title"]
        for row in conn.execute("SELECT item_no, title FROM legal_item")
    }


def ensure_tacit_knowledge_schema(conn: sqlite3.Connection) -> None:
    expected_columns = ["item_no", "title", "content", "created_at"]
    columns = [row[1] for row in conn.execute("PRAGMA table_info(legal_tacit_knowledge)")]
    if columns == expected_columns:
        return

    order_by = []
    if "priority" in columns:
        order_by.append("priority")
    if "id" in columns:
        order_by.append("id")
    order_clause = f" ORDER BY {', '.join(order_by)}" if order_by else ""

    conn.execute("ALTER TABLE legal_tacit_knowledge RENAME TO legal_tacit_knowledge_old")
    conn.execute(
        """
        CREATE TABLE legal_tacit_knowledge (
            item_no TEXT NOT NULL,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        f"""
        INSERT INTO legal_tacit_knowledge (item_no, title, content, created_at)
        SELECT item_no, title, content, COALESCE(created_at, CURRENT_TIMESTAMP)
        FROM legal_tacit_knowledge_old{order_clause}
        """
    )
    conn.execute("DROP TABLE legal_tacit_knowledge_old")


def load_tacit_knowledge_hits(conn: sqlite3.Connection, item_no: str) -> list[RagHit]:
    if not table_exists(conn, "legal_tacit_knowledge"):
        return []
    item_group = item_no_group(item_no) + ["공통"]
    if not item_group:
        return []
    placeholders = ", ".join("?" for _ in item_group)
    rows = conn.execute(
        f"""
        SELECT item_no, title, content
        FROM legal_tacit_knowledge
        WHERE item_no IN ({placeholders})
        ORDER BY rowid
        """,
        item_group,
    ).fetchall()
    return [
        RagHit(
            item_no=row["item_no"],
            source_type="tacit_knowledge",
            source_name="legal_tacit_knowledge",
            title=row["title"] or "",
            category="operational_tacit_knowledge",
            snippet=row["content"] or "",
            page_or_row=None,
            score=None,
        )
        for row in rows
    ]


def load_authoritative_criteria_hits(conn: sqlite3.Connection, item_no: str) -> list[RagHit]:
    if not table_exists(conn, "legal_item"):
        return []
    item_group = item_no_group(item_no)
    if not item_group:
        return []
    placeholders = ", ".join("?" for _ in item_group)
    requirements_by_item = load_requirements_by_item(conn)
    reference_examples_by_item = load_reference_examples_by_item(conn)
    rows = conn.execute(
        f"""
        SELECT item_no, title, target_text
        FROM legal_item
        WHERE item_no IN ({placeholders})
        ORDER BY item_no
        """,
        item_group,
    ).fetchall()
    return [
        RagHit(
            item_no=row["item_no"],
            source_type="authoritative_criteria_db",
            source_name=source_name_for_criteria(reference_examples_by_item.get(row["item_no"], [])),
            title=row["title"] or "",
            category="legal_review_criteria",
            snippet=build_enriched_criteria_content(
                row,
                requirements_by_item.get(row["item_no"], []),
                reference_examples_by_item.get(row["item_no"], []),
            ),
            page_or_row=None,
            score=None,
        )
        for row in rows
    ]


def merge_rag_hits(primary: list[RagHit], secondary: list[RagHit]) -> list[RagHit]:
    result: list[RagHit] = []
    seen: set[tuple[str | None, str, str, str]] = set()
    for hit in primary + secondary:
        key = (hit.item_no, hit.source_type, hit.title, hit.snippet[:160])
        if key in seen:
            continue
        seen.add(key)
        result.append(hit)
    return result


def build_enriched_criteria_content(
    row: sqlite3.Row,
    requirements: list[dict[str, str]],
    reference_examples: list[dict[str, str]],
) -> str:
    sections = [
        "[TARGET]",
        f"item_no: {row['item_no']}",
        f"legal_title: {row['title']}",
        f"target_text: {row['target_text']}" if row["target_text"] else "",
    ]
    sections.append("[REQUIRED_CRITERIA]")
    if requirements:
        for idx, requirement in enumerate(requirements, start=1):
            sections.append(format_requirement_block(idx, requirement))
    else:
        sections.append("no legal_requirement rows")

    if str(row["item_no"]) == "14" and reference_examples:
        sections.append("[REFERENCE_EXAMPLES_FOR_ITEM_14]")
        for idx, example in enumerate(reference_examples, start=1):
            sections.append(format_reference_example_block(idx, example))

    return "\n".join(section for section in sections if section)


def build_requirement_content(requirement: dict[str, str]) -> str:
    core_requirement_text = core_requirement(requirement)
    return "\n".join(
        section
        for section in [
            f"category: {requirement['category']}" if requirement["category"] else "",
            f"core_requirement_text: {core_requirement_text}" if core_requirement_text else "",
            f"requirement_text: {requirement['requirement_text']}",
            f"example_sentence: {requirement['example_sentence']}"
            if requirement["example_sentence"]
            else "",
            f"case_text: {requirement['case_text']}" if requirement["case_text"] else "",
        ]
        if section
    )


def format_requirement_block(idx: int, requirement: dict[str, str]) -> str:
    core_requirement_text = core_requirement(requirement)
    lines = [
        f"[CRITERION {idx}]",
        f"category: {requirement['category'] or 'criteria'}",
        f"core_requirement_text: {core_requirement_text}",
        f"requirement_text: {requirement['requirement_text']}",
    ]
    if requirement["example_sentence"]:
        lines.append("example_sentence_note: reference example only; exact wording is not mandatory")
        lines.append(f"example_sentence: {requirement['example_sentence']}")
    if requirement["case_text"]:
        lines.append("case_text_note: conditional criterion; apply only if this RFP matches the case")
        lines.append(f"case_text: {requirement['case_text']}")
    return "\n".join(lines)


def core_requirement(requirement: dict[str, str]) -> str:
    text = requirement["requirement_text"].strip()
    example = requirement["example_sentence"].strip()
    if not example:
        return text
    normalized_text = normalize_space(text)
    normalized_example = normalize_space(example)
    if normalized_example and normalized_example in normalized_text:
        return normalize_space(normalized_text.replace(normalized_example, " ")).strip()
    return text


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def format_reference_example_block(idx: int, example: dict[str, str]) -> str:
    lines = [
        f"[REFERENCE_EXAMPLE {idx}]",
        f"reference_type: {example['reference_type'] or 'reference_example'}",
    ]
    if example["reference_subtype"]:
        lines.append(f"reference_subtype: {example['reference_subtype']}")
    lines.append(f"content: {example['content']}")
    return "\n".join(lines)


def source_name_for_criteria(reference_examples: list[dict[str, str]]) -> str:
    if reference_examples:
        return "legal_item+legal_requirement+legal_reference_example"
    return "legal_item+legal_requirement"


def criteria_source_tables(reference_examples: list[dict[str, str]]) -> list[str]:
    tables = ["legal_item", "legal_requirement"]
    if reference_examples:
        tables.append("legal_reference_example")
    return tables


def make_requirement_title(
    item_no: str,
    item_title: str | None,
    category: str | None,
    requirement_text: str,
) -> str:
    base = item_title or f"{item_no} legal item"
    label = category or "criteria"
    first_line = compact_cell((requirement_text or "").splitlines()[0])
    if first_line.startswith("ㅁ "):
        first_line = first_line[2:]
    return f"{base} - {label}: {first_line[:80]}"


def insert_rag_document(conn: sqlite3.Connection, doc: dict[str, object]) -> None:
    cur = conn.execute(
        """
        INSERT INTO rag_document (
            source_type, source_name, sheet_name, row_no, item_no,
            title, category, content, metadata_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            doc["source_type"],
            doc["source_name"],
            doc.get("sheet_name"),
            doc.get("row_no"),
            doc.get("item_no"),
            doc.get("title"),
            doc.get("category"),
            doc.get("content") or "",
            json.dumps(doc.get("metadata", {}), ensure_ascii=False),
        ),
    )
    conn.execute(
        """
        INSERT INTO rag_fts (
            title, category, content, item_no, source_type, source_name, sheet_name, row_no, doc_id
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            doc.get("title") or "",
            doc.get("category") or "",
            doc.get("content") or "",
            doc.get("item_no") or "",
            doc["source_type"],
            doc["source_name"],
            doc.get("sheet_name"),
            doc.get("row_no"),
            cur.lastrowid,
        ),
    )


def search(conn: sqlite3.Connection, query: str, item_no: str | None, limit: int) -> list[sqlite3.Row]:
    params: list[object] = [make_match_query(query)]
    where = "rag_fts MATCH ?"
    item_group = item_no_group(item_no)
    if item_group:
        placeholders = ", ".join("?" for _ in item_group)
        where += f" AND item_no IN ({placeholders})"
        params.extend(item_group)
    params.append(limit)

    rows = conn.execute(
        f"""
        SELECT
            doc_id, item_no, source_type, source_name, sheet_name, row_no,
            title, category, content AS snippet, bm25(rag_fts) AS rank
        FROM rag_fts
        WHERE {where}
        ORDER BY
            CASE source_type
                WHEN 'criteria_db' THEN 1
                WHEN 'requirement_db' THEN 2
                WHEN 'reference_example_db' THEN 3
                ELSE 4
            END,
            rank
        LIMIT ?
        """,
        params,
    ).fetchall()
    if rows:
        return rows
    return like_search(conn, query, item_no, limit)


def item_document_fallback(conn: sqlite3.Connection, item_no: str | None, limit: int) -> list[sqlite3.Row]:
    item_group = item_no_group(item_no)
    if not item_group:
        return []
    placeholders = ", ".join("?" for _ in item_group)
    return conn.execute(
        f"""
        SELECT
            id AS doc_id, item_no, source_type, source_name, sheet_name, row_no,
            title, category, content AS snippet, 0 AS rank
        FROM rag_document
        WHERE item_no IN ({placeholders})
        ORDER BY
            CASE source_type
                WHEN 'requirement_db' THEN 1
                WHEN 'criteria_db' THEN 2
                WHEN 'reference_example_db' THEN 3
                ELSE 4
            END,
            row_no,
            id
        LIMIT ?
        """,
        item_group + [limit],
    ).fetchall()


def like_search(conn: sqlite3.Connection, query: str, item_no: str | None, limit: int) -> list[sqlite3.Row]:
    tokens = TOKEN_RE.findall(query)
    clauses = []
    params: list[object] = []
    for token in tokens[:5]:
        clauses.append("(title LIKE ? OR category LIKE ? OR content LIKE ?)")
        like = f"%{token}%"
        params.extend([like, like, like])
    where = " OR ".join(clauses) if clauses else "1 = 1"
    item_group = item_no_group(item_no)
    if item_group:
        placeholders = ", ".join("?" for _ in item_group)
        where = f"({where}) AND item_no IN ({placeholders})"
        params.extend(item_group)
    params.append(limit)
    return conn.execute(
        f"""
        SELECT
            id AS doc_id, item_no, source_type, source_name, sheet_name, row_no,
            title, category, content AS snippet, 0 AS rank
        FROM rag_document
        WHERE {where}
        ORDER BY
            CASE source_type
                WHEN 'criteria_db' THEN 1
                WHEN 'requirement_db' THEN 2
                WHEN 'reference_example_db' THEN 3
                ELSE 4
            END,
            row_no
        LIMIT ?
        """,
        params,
    ).fetchall()


def documents_from_excel(path: Path) -> list[dict[str, object]]:
    if load_workbook is None:
        return []
    docs: list[dict[str, object]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = [compact_cell(value) for value in row]
                filled = [cell for cell in cells if cell]
                if not filled:
                    continue
                item_no = extract_item_no(cells[:5])
                docs.append(
                    {
                        "source_type": "criteria_excel",
                        "source_name": path.name,
                        "sheet_name": ws.title,
                        "row_no": row_no,
                        "item_no": item_no,
                        "title": filled[0][:120],
                        "category": "criteria_excel",
                        "content": "\n".join(
                            f"{idx + 1}: {cell}" for idx, cell in enumerate(cells) if cell
                        ),
                        "metadata": {"excel_file": path.name, "sheet": ws.title, "row": row_no},
                    }
                )
    finally:
        wb.close()
    return docs


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    return (
        conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
            (table_name,),
        ).fetchone()
        is not None
    )


def item_no_group(item_no: str | None) -> list[str]:
    if not item_no:
        return []
    normalized = str(item_no).strip()
    return ITEM_ALIASES.get(normalized, [normalized])


def default_query_for_item(item_no: str) -> str:
    group = item_no_group(item_no)
    return " ".join(group + ["legal", "criteria", "requirement", "review"])


def context_hit_limit(item_no: str | None) -> int:
    if str(item_no or "").strip() in {"2", "2-1"}:
        return 2
    return 1


def make_match_query(query: str) -> str:
    tokens = TOKEN_RE.findall(query)
    if not tokens:
        return '""'
    return " OR ".join(f'"{token}"' for token in tokens[:12])


def compact_cell(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def extract_item_no(values: Iterable[str]) -> str | None:
    for value in values:
        match = re.search(r"(?<!\d)(\d{1,2}(?:-\d{1,2})?)(?!\d)", value or "")
        if match:
            return match.group(1)
    return None


def format_location(row: sqlite3.Row) -> str | None:
    parts = []
    if row["sheet_name"]:
        parts.append(str(row["sheet_name"]))
    if row["row_no"]:
        parts.append(f"row {row['row_no']}")
    return " / ".join(parts) if parts else None
