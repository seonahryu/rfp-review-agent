from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from agents.gpt_judgement import is_target_only_evidence
from agents.models import AuditWarning, FinalReview


ITEM_NAMES = {
    "1": "과업심의위원회",
    "2": "상용SW 직접구매 및 SW품질성능 평가시험(BMT)",
    "2-1": "상용SW 직접구매 제외사유",
    "3": "중소 SW사업자의 사업 참여 지원",
    "4": "하도급 제한",
    "5": "SW사업 작업장소(원격개발)",
    "6": "SW사업 산출물 활용 보장",
    "7": "개발SW의 공동활용 사전명시",
    "8": "하자담보 책임기간 및 범위",
    "9": "특정규격 명시 금지",
    "10": "협상에 의한 계약 우선 적용",
    "11": "기술능력 평가비중",
    "12": "SW기술성 평가기준 적용",
    "13": "SW사업 제안서 보상",
    "14": "요구사항 상세화",
    "15": "작업장소 보안 및 개인정보보호",
    "16": "투입인력 요구 및 관리 금지",
    "17": "SW사업 영향평가",
    "18": "SW사업정보 제출",
}


class ReportAgent:
    def __init__(self, output_dir: Path | str = "outputs") -> None:
        self.output_dir = Path(output_dir)

    def write_excel(
        self,
        document_id: int,
        final_reviews: list[FinalReview],
        audit_warnings: list[AuditWarning],
    ) -> Path:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        path = self.output_dir / f"rfp_review_document_{document_id}.xlsx"

        wb = Workbook()
        compliance = wb.active
        compliance.title = "붙임1_법령준수여부"
        compliance.append(["검토항목", "법령준수 여부", "개선권고 관련 법적 근거", "사람 재검토"])
        for review in final_reviews:
            compliance.append(
                [
                    item_label(review.item_no),
                    compliance_label(review),
                    basis_text(review),
                    human_review_label(review),
                ]
            )

        detail = wb.create_sheet("별첨_항목별검토")
        detail.append(["항목", "권고내용", ""])
        for review in final_reviews:
            detail.append([f"[{compliance_label(review)}] {item_label(review.item_no)}", copy_paste_comment(review), ""])

        opinion = wb.create_sheet("법제도_검토의견")
        write_opinion_sheet(opinion, final_reviews)

        for sheet in wb.worksheets:
            style_sheet(sheet)
        wb.save(path)
        return path


def item_label(item_no: str) -> str:
    return f"{item_no}. {ITEM_NAMES.get(item_no, '검토항목')}"


def compliance_label(review: FinalReview) -> str:
    if str(review.item_no) == "16":
        if review.is_target is True:
            return "대상"
        if review.is_target is False:
            return "비대상"
        return "대상여부 확인필요"
    result = (review.final_result or "").strip()
    if result in {"준수", "미준수", "보완필요"}:
        return result
    if result in {"비대상", "해당없음"} or review.is_target is False:
        return "해당없음"
    if review.final_status == "자동 확정 가능" and result:
        return result
    return "보완필요"


def human_review_label(review: FinalReview) -> str:
    if review.final_status == "자동 확정 가능" and not review.warnings:
        return "아니오"
    return "예"


def basis_text(review: FinalReview) -> str:
    if str(review.item_no) == "9" and compliance_label(review) == "준수":
        return (
            "판단근거: 특정규격 명시 없음. RFP 본문에서 특정 상표, 모델명, "
            "특정 제품에 종속된 기술용어가 확인되지 않아 해당 기준을 충족하는 것으로 판단됩니다."
        )

    parts = []
    if review.recommendation:
        parts.append(f"개선권고: {review.recommendation}")
    if review.reason:
        parts.append(f"판단근거: {review.reason}")
    evidence = evidence_trigger_text(review)
    if evidence:
        parts.append(f"RFP 트리거: {evidence}")
    if not parts:
        return "근거 확인 필요"
    return "\n".join(parts)


def evidence_trigger_text(review: FinalReview) -> str:
    lines = []
    pages, texts = reportable_evidence(review)
    for idx, text in enumerate(texts[:5]):
        page_label = ""
        if idx < len(pages):
            page_label = f"p.{pages[idx]} "
        snippet = compact_evidence(text)
        if snippet:
            lines.append(f"{page_label}{snippet}".strip())
    if lines:
        return "\n".join(lines)
    if pages:
        return ", ".join(f"p.{page}" for page in pages[:5])
    return ""


def compact_evidence(text: str, limit: int = 260) -> str:
    import re

    compacted = re.sub(r"\s+", " ", str(text or "")).strip()
    if len(compacted) <= limit:
        return compacted
    return compacted[: limit - 3].rstrip() + "..."


def copy_paste_comment(review: FinalReview) -> str:
    pages = format_pages(reportable_evidence_pages(review))
    label = compliance_label(review)
    generated_content = getattr(getattr(review, "compliance_content", None), "compliance_content", "")
    if generated_content and label == "준수":
        return generated_content
    if str(review.item_no) == "16":
        if review.is_target is True:
            return f"제안요청서 {pages} 기준 대상 사업으로 판단" if pages else "대상 사업으로 판단"
        if review.is_target is False:
            return f"제안요청서 {pages} 기준 비대상 사업으로 판단" if pages else "비대상 사업으로 판단"
        return "대상여부 확인필요"
    if label == "준수":
        if str(review.item_no) == "5":
            return detail_comment_item_5(review)
        if str(review.item_no) == "6":
            return detail_comment_item_6(review)
        if str(review.item_no) == "9":
            return "특정규격 명시 없음"
        if str(review.item_no) == "14" and review.evidence_pages:
            first_page = min(review.evidence_pages)
            return f"제안요청서 p.{first_page} 이하 명시"
        return f"제안요청서 {pages} 명시" if pages else "관련 문구 명시"
    if label == "해당없음":
        return "해당없음"
    if label in {"미준수", "보완필요"}:
        recommendation = review.recommendation or review.reason or "관련 법령 기준에 맞게 보완하시기 바랍니다."
        if str(review.item_no) in {"12", "15", "17"} and review.recommendation:
            return review.recommendation
        return f"제안요청서 {pages} 일부 명시\n→ {recommendation}" if pages else recommendation
    reason = review.reason or "자동 판정 근거가 부족합니다."
    return f"확인필요 - {reason}"


def detail_comment_item_5(review: FinalReview) -> str:
    text = joined_evidence(review)
    rows = [
        (
            "1",
            "<작업장소 상호협의 등>",
            "작업장소 상호협의 또는 제공여부 및 관련 비용 계상여부",
            has_any(text, ["작업장소 등은 상호 협의", "작업장소 상호협의", "사업예산 내 계상", "관련 비용을 포함"]),
        ),
        (
            "2",
            "<원격지 개발 장소 제시·검토 절차>",
            "공급자에 의한 작업장소 제시 및 발주기관의 검토절차",
            has_any(text, ["작업장소를 제시", "공급자 제시", "제시된 작업장소", "개발방법 등에 대한 구체적인 방안"]),
        ),
        (
            "3",
            "<원격지 개발 장소 보안요구사항>",
            "작업장소 관련 보안요구사항",
            has_any(text, ["원격지 보안관리", "보안요구사항", "보안사고", "출입보안", "출입통제"]),
        ),
    ]
    return detail_table_comment(review, rows)


def detail_comment_item_6(review: FinalReview) -> str:
    text = joined_evidence(review)
    rows = [
        (
            "1",
            "<지식재산권 공동귀속>",
            "제안요청서 등에 지식재산의 공동귀속 적용 여부",
            has_any(text, ["지식재산권 귀속", "지식재산 공동귀속", "공동으로 소유", "용역계약일반조건"]),
        ),
        (
            "2",
            "<SW산출물 반출 절차 등>",
            "SW산출물 반출 요청절차",
            has_any(text, ["산출물의 반출", "산출물 반출", "반출을 요청"]),
        ),
        (
            "2-1",
            "<SW산출물 반출 절차 등>",
            "누출금지정보 삭제 및 확약서 제출",
            has_any(text, ["누출금지정보", "누출금지 정보"]) and has_any(text, ["확약서", "삭제"]),
        ),
        (
            "2-2",
            "<SW산출물 반출 절차 등>",
            "제3자 제공 시 발주기관 사전승인",
            has_any(text, ["제3자", "제 3자", "제3자에게"]) and has_any(text, ["사전승인", "사전 승인"]),
        ),
        (
            "2-3",
            "<SW산출물 반출 절차 등>",
            "무단 유출/누출금지정보 미삭제 시 입찰참가자격 제한",
            has_any(text, ["무단 유출", "무단으로 유출", "누출되는 경우"]) and has_any(
                text, ["입찰 참가 자격", "입찰참가자격", "입찰 참가 자격을 제한"]
            ),
        ),
    ]
    return detail_table_comment(review, rows)


def detail_table_comment(review: FinalReview, rows: list[tuple[str, str, str, bool]]) -> str:
    page_text = format_pages(review.evidence_pages)
    all_present = all(is_present for _, _, _, is_present in rows)
    if all_present:
        return f"제안요청서 {page_text} 명시" if page_text else "관련 문구 명시"
    missing = ", ".join(description for _, _, description, is_present in rows if not is_present)
    return (
        f"제안요청서 {page_text} 일부 미명시 - {missing}"
        if page_text
        else f"일부 미명시 - {missing}"
    )


def joined_evidence(review: FinalReview) -> str:
    return " ".join(str(text or "") for text in reportable_evidence(review)[1])


def reportable_evidence_pages(review: FinalReview) -> list[int]:
    return reportable_evidence(review)[0]


def reportable_evidence(review: FinalReview) -> tuple[list[int], list[str]]:
    if compliance_label(review) == "해당없음" or str(review.item_no) == "16":
        return list(review.evidence_pages), list(review.evidence_text)
    pages: list[int] = []
    texts: list[str] = []
    for idx, page in enumerate(review.evidence_pages):
        text = review.evidence_text[idx] if idx < len(review.evidence_text) else ""
        if text and is_target_only_evidence(text):
            continue
        pages.append(page)
        texts.append(text)
    return pages, texts


def has_any(text: str, terms: list[str]) -> bool:
    compact = normalized_for_match(text)
    return any(normalized_for_match(term) in compact for term in terms)


def normalized_for_match(text: str) -> str:
    return "".join(str(text or "").split())


def write_opinion_sheet(ws, final_reviews: list[FinalReview]) -> None:
    non_compliant = [review for review in final_reviews if compliance_label(review) == "미준수"]
    recommendations = [review for review in final_reviews if compliance_label(review) == "보완필요"]
    ws.append(
        [
            f"총 {len(final_reviews)}개 항목 중 {len(non_compliant)}개 항목 미준수 및 "
            f"{len(recommendations)}개 항목 보완 권고"
        ]
    )
    ws.append([""])
    for review in non_compliant + recommendations:
        ws.append([item_label(review.item_no)])
        ws.append([f" - {review.recommendation or review.reason or '관련 기준에 맞게 보완 필요'}"])
        ws.append([""])


def format_pages(pages: list[int]) -> str:
    if not pages:
        return ""
    unique = sorted(set(pages))
    if len(unique) == 1:
        return f"p.{unique[0]}"
    ranges = []
    start = previous = unique[0]
    for page in unique[1:]:
        if page == previous + 1:
            previous = page
            continue
        ranges.append(format_page_range(start, previous))
        start = previous = page
    ranges.append(format_page_range(start, previous))
    return "pp." + ", ".join(ranges[:5])


def format_page_range(start: int, end: int) -> str:
    if start == end:
        return str(start)
    return f"{start}-{end}"


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 80)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
